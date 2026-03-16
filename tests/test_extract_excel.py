"""Step2 Excel 構造抽出テスト

確認観点:
  - シート名が見出し要素として抽出される
  - セルテキストが正しく読み取れる
  - 結合セルが検出・処理される
  - 非表示シートがスキップされる
  - コメントがセルテキストに付記される
  - 空シートが処理される
  - 大きなシートで警告が出る
  - 複数シートが全て抽出される
  - 存在しないファイルでエラーが返る
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from src.config import PipelineConfig
from src.extractors.excel import extract_xlsx
from src.models.metadata import ProcessStatus


class TestBasicExtraction:
    """基本的な抽出テスト"""

    def test_simple_table(self, simple_xlsx: Path, config: PipelineConfig):
        """基本的なデータテーブルが抽出されること"""
        record, result = extract_xlsx(simple_xlsx, "simple.xlsx", ".xlsx", config)

        assert result.status in (ProcessStatus.SUCCESS, ProcessStatus.WARNING)
        elements = record.document["elements"]

        # シート名が見出しとして抽出
        headings = [e for e in elements if e["type"] == "heading"]
        assert len(headings) >= 1
        assert headings[0]["content"]["text"] == "エラーコード一覧"
        assert headings[0]["content"]["detection_method"] == "sheet_name"

        # テーブルが抽出される
        tables = [e for e in elements if e["type"] == "table"]
        assert len(tables) == 1

        rows = tables[0]["content"]["rows"]
        assert len(rows) == 4  # ヘッダー + 3データ行

        # ヘッダー行
        header_texts = [cell["text"] for cell in rows[0]]
        assert "コード" in header_texts
        assert "メッセージ" in header_texts

        # データ行
        first_data_texts = [cell["text"] for cell in rows[1]]
        assert "E001" in first_data_texts
        assert "入力値が不正です" in first_data_texts

    def test_metadata(self, simple_xlsx: Path, config: PipelineConfig):
        """メタデータが正しく設定されること"""
        record, result = extract_xlsx(simple_xlsx, "simple.xlsx", ".xlsx", config)

        assert record.metadata.source_path == "simple.xlsx"
        assert record.metadata.source_ext == ".xlsx"
        assert record.metadata.source_size_bytes > 0
        assert record.metadata.doc_role_guess == "data_sheet"


class TestMultiSheet:
    """複数シートのテスト"""

    def test_all_sheets_extracted(self, multi_sheet_xlsx: Path, config: PipelineConfig):
        """全シートが抽出されること"""
        record, result = extract_xlsx(
            multi_sheet_xlsx, "multi.xlsx", ".xlsx", config,
        )

        elements = record.document["elements"]

        headings = [e for e in elements if e["type"] == "heading"]
        heading_texts = [h["content"]["text"] for h in headings]
        assert "機能一覧" in heading_texts
        assert "エラーコード" in heading_texts

        tables = [e for e in elements if e["type"] == "table"]
        assert len(tables) == 2

    def test_sheet_order_preserved(self, multi_sheet_xlsx: Path, config: PipelineConfig):
        """シートの出現順序が保持されること"""
        record, _ = extract_xlsx(multi_sheet_xlsx, "multi.xlsx", ".xlsx", config)

        elements = record.document["elements"]
        headings = [e for e in elements if e["type"] == "heading"]
        assert headings[0]["content"]["text"] == "機能一覧"
        assert headings[1]["content"]["text"] == "エラーコード"


class TestMergedCells:
    """結合セルのテスト"""

    def test_merged_cells_detected(self, merged_cells_xlsx: Path, config: PipelineConfig):
        """結合セルが検出されること"""
        record, result = extract_xlsx(
            merged_cells_xlsx, "merged.xlsx", ".xlsx", config,
        )

        elements = record.document["elements"]
        tables = [e for e in elements if e["type"] == "table"]
        assert len(tables) >= 1

        table = tables[0]["content"]
        assert table["has_merged_cells"] is True
        assert table["confidence"] == "medium"

    def test_merged_cell_text(self, merged_cells_xlsx: Path, config: PipelineConfig):
        """結合セルのテキストが正しく取得されること"""
        record, _ = extract_xlsx(
            merged_cells_xlsx, "merged.xlsx", ".xlsx", config,
        )

        elements = record.document["elements"]
        tables = [e for e in elements if e["type"] == "table"]
        table_rows = tables[0]["content"]["rows"]

        # 全セルのテキストをフラット化
        all_texts = [
            cell["text"] for row in table_rows for cell in row if cell["text"]
        ]
        assert "接続設定" in all_texts
        assert "ホスト名" in all_texts
        assert "db-server01" in all_texts

    def test_horizontal_merge_colspan(self, merged_cells_xlsx: Path, config: PipelineConfig):
        """横結合セルの colspan が正しく設定されること"""
        record, _ = extract_xlsx(
            merged_cells_xlsx, "merged.xlsx", ".xlsx", config,
        )

        elements = record.document["elements"]
        tables = [e for e in elements if e["type"] == "table"]
        table_rows = tables[0]["content"]["rows"]

        # A4:C4 の横結合セルを探す
        banner_cells = [
            cell for row in table_rows for cell in row
            if cell.get("colspan", 1) == 3
        ]
        assert len(banner_cells) >= 1
        assert "認証設定" in banner_cells[0]["text"]


class TestHiddenSheet:
    """非表示シートのテスト"""

    def test_hidden_sheet_skipped(self, hidden_sheet_xlsx: Path, config: PipelineConfig):
        """非表示シートがスキップされること"""
        record, result = extract_xlsx(
            hidden_sheet_xlsx, "hidden.xlsx", ".xlsx", config,
        )

        elements = record.document["elements"]
        headings = [e for e in elements if e["type"] == "heading"]
        heading_texts = [h["content"]["text"] for h in headings]

        assert "表示シート" in heading_texts
        assert "マスタ" not in heading_texts

    def test_hidden_sheet_warning(self, hidden_sheet_xlsx: Path, config: PipelineConfig):
        """非表示シートがあれば警告が出ること"""
        _, result = extract_xlsx(
            hidden_sheet_xlsx, "hidden.xlsx", ".xlsx", config,
        )

        assert result.status == ProcessStatus.WARNING
        assert "hidden_sheets=" in result.message


class TestComments:
    """コメントのテスト"""

    def test_comment_appended_to_text(self, comments_xlsx: Path, config: PipelineConfig):
        """コメントがセルテキストに付記されること"""
        record, _ = extract_xlsx(comments_xlsx, "comments.xlsx", ".xlsx", config)

        elements = record.document["elements"]
        tables = [e for e in elements if e["type"] == "table"]
        table_rows = tables[0]["content"]["rows"]

        # B2 に "C001 ※注: ..." が付くはず
        all_texts = [cell["text"] for row in table_rows for cell in row]
        comment_texts = [t for t in all_texts if "※注:" in t]
        assert len(comment_texts) >= 1
        assert any("外部IF" in t for t in comment_texts)


class TestLayoutSegmentation:
    """レイアウト分割のテスト"""

    def test_separated_regions_become_multiple_tables(self, tmp_path: Path, config: PipelineConfig):
        """空白行・空白列で分離された領域が別テーブルとして抽出されること"""
        wb = Workbook()
        ws = wb.active
        ws.title = "複数領域"

        ws["A1"] = "帳票一覧"
        ws.merge_cells("A1:B1")
        ws["A3"] = "ID"
        ws["B3"] = "名称"
        ws["A4"] = "R1"
        ws["B4"] = "売上日報"

        ws["E1"] = "バッチ一覧"
        ws.merge_cells("E1:F1")
        ws["E3"] = "ID"
        ws["F3"] = "処理名"
        ws["E4"] = "B1"
        ws["F4"] = "日次集計"

        path = tmp_path / "segmented.xlsx"
        wb.save(path)

        record, _ = extract_xlsx(path, "segmented.xlsx", ".xlsx", config)
        tables = [e for e in record.document["elements"] if e["type"] == "table"]

        assert len(tables) >= 4
        flattened = [
            [cell["text"] for row in table["content"]["rows"] for cell in row]
            for table in tables
        ]
        assert any("帳票一覧" in texts for texts in flattened)
        assert any("バッチ一覧" in texts for texts in flattened)
        assert any("売上日報" in texts for texts in flattened)
        assert any("日次集計" in texts for texts in flattened)


class TestFormulaFallback:
    """数式セルのフォールバックテスト"""

    def test_formula_text_used_when_cached_value_missing(self, tmp_path: Path, config: PipelineConfig):
        """data_only で値が空でも数式文字列が残ること"""
        wb = Workbook()
        ws = wb.active
        ws.title = "集計"
        ws.append(["項目", "値"])
        ws.append(["売上", 100])
        ws.append(["合計", "=SUM(B2:B2)"])

        path = tmp_path / "formula.xlsx"
        wb.save(path)

        record, _ = extract_xlsx(path, "formula.xlsx", ".xlsx", config)
        tables = [e for e in record.document["elements"] if e["type"] == "table"]
        all_texts = [
            cell["text"]
            for table in tables
            for row in table["content"]["rows"]
            for cell in row
        ]

        assert "=SUM(B2:B2)" in all_texts


class TestEmptySheet:
    """空シートのテスト"""

    def test_empty_sheet_handled(self, empty_xlsx: Path, config: PipelineConfig):
        """空のシートでもエラーにならないこと"""
        record, result = extract_xlsx(empty_xlsx, "empty.xlsx", ".xlsx", config)

        assert result.status in (ProcessStatus.SUCCESS, ProcessStatus.WARNING)
        elements = record.document["elements"]

        # シート名の見出しは出る
        headings = [e for e in elements if e["type"] == "heading"]
        assert len(headings) >= 1

        # テーブルは出ない
        tables = [e for e in elements if e["type"] == "table"]
        assert len(tables) == 0


class TestLargeSheet:
    """大きなシートのテスト"""

    def test_large_sheet_warning(self, large_xlsx: Path, config: PipelineConfig):
        """大きなシートで警告が出ること"""
        _, result = extract_xlsx(large_xlsx, "large.xlsx", ".xlsx", config)

        assert result.status == ProcessStatus.WARNING
        assert "large_sheet:" in result.message


class TestErrorHandling:
    """エラーハンドリングテスト"""

    def test_nonexistent_file(self, tmp_path: Path, config: PipelineConfig):
        """存在しないファイルでエラーが返ること"""
        fake_path = tmp_path / "nonexistent.xlsx"
        record, result = extract_xlsx(fake_path, "nonexistent.xlsx", ".xlsx", config)

        assert result.status == ProcessStatus.ERROR
        assert record.document == {}


class TestRegistryIntegration:
    """レジストリとの統合テスト"""

    def test_xlsx_registered(self):
        """xlsx 拡張子がレジストリに登録されていること"""
        from src.extractors.registry import get_extractor, supported_extensions

        assert ".xlsx" in supported_extensions()
        extractor = get_extractor(".xlsx")
        assert extractor is not None


class TestMarkdownTransform:
    """Excel → 中間表現 → Markdown 変換の結合テスト"""

    def test_end_to_end(self, simple_xlsx: Path, config: PipelineConfig):
        """抽出した中間表現が Markdown に変換できること"""
        from src.transform.to_markdown import transform_to_markdown

        record, _ = extract_xlsx(simple_xlsx, "simple.xlsx", ".xlsx", config)
        md = transform_to_markdown(record.to_dict())

        assert "エラーコード一覧" in md
        assert "E001" in md
        assert "入力値が不正です" in md
        # 半構造化テキストとして「コード:」ラベルが出る
        assert "コード:" in md or "コード" in md
