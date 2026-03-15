"""xlsx 詳細調査スクリプト

.xlsx / .xlsm ファイルの内部構造を詳細に調査し、パイプライン設計に必要な
情報を抽出する。profile_documents.py より深い情報を取る。

使い方:
  # 単一ファイルを調査
  python tools/inspect_xlsx.py <file.xlsx> -o inspect_xlsx_report

  # フォルダ配下の全 .xlsx / .xlsm を調査
  python tools/inspect_xlsx.py <folder> -o inspect_xlsx_report

出力:
  inspect_xlsx_report.json  - 詳細データ
  inspect_xlsx_report.txt   - 口頭説明用テキスト
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from collections import Counter
from dataclasses import asdict, dataclass, field
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("エラー: openpyxl がインストールされていません (pip install openpyxl)", file=sys.stderr)
    sys.exit(1)


FORMAT_SAMPLE_ROW_LIMIT = 1000
MERGED_SAMPLE_LIMIT = 10
COMMENT_SAMPLE_LIMIT = 5
NUMBER_FORMAT_TOP_N = 10
COLOR_PATTERN_TOP_N = 10
LARGE_SHEET_ROW_THRESHOLD = 100
LARGE_SHEET_COL_THRESHOLD = 20
HIGH_EMPTY_RATIO_THRESHOLD = 0.8
FORM_GRID_STDDEV_THRESHOLD = 0.5


# ---------------------------------------------------------------------------
# データクラス
# ---------------------------------------------------------------------------
@dataclass
class DimensionStats:
    min: float
    max: float
    avg: float
    stddev: float


@dataclass
class SheetDetail:
    name: str
    index: int
    is_hidden: bool
    is_protected: bool
    max_row: int
    max_column: int
    non_empty_cells: int
    total_cells: int
    empty_ratio: float
    has_merged_cells: bool
    merged_cell_count: int
    merged_cell_ranges: list[str] = field(default_factory=list)
    has_tables: bool = False
    table_count: int = 0
    table_names: list[str] = field(default_factory=list)
    has_images: bool = False
    image_count: int = 0
    has_comments: bool = False
    comment_count: int = 0
    comment_samples: list[str] = field(default_factory=list)
    has_formulas: bool = False
    formula_count: int = 0
    has_conditional_formatting: bool = False
    cf_rule_count: int = 0
    has_data_validation: bool = False
    dv_count: int = 0
    has_auto_filter: bool = False
    filter_ref: str | None = None
    has_row_outline: bool = False
    has_col_outline: bool = False
    has_hidden_rows: bool = False
    hidden_row_count: int = 0
    has_hidden_cols: bool = False
    hidden_col_count: int = 0
    has_freeze_panes: bool = False
    freeze_panes_ref: str | None = None
    print_area: str | None = None
    print_title_rows: str | None = None
    print_title_cols: str | None = None
    number_format_patterns: dict[str, int] = field(default_factory=dict)
    color_fill_patterns: dict[str, int] = field(default_factory=dict)
    grid_lines_visible: bool = True
    column_width_stats: dict[str, float] | None = None
    row_height_stats: dict[str, float] | None = None


@dataclass
class NamedRangeInfo:
    name: str
    scope: str
    refers_to: str


@dataclass
class XlsxInspection:
    path: str
    error: str | None = None
    sheet_count: int = 0
    hidden_sheet_count: int = 0
    sheets: list[dict] = field(default_factory=list)
    named_ranges: list[dict] = field(default_factory=list)
    has_vba: bool = False
    total_merged_cells: int = 0
    total_formulas: int = 0
    total_comments: int = 0
    total_images: int = 0
    estimated_layout_type: str = "unknown"


# ---------------------------------------------------------------------------
# ヘルパー
# ---------------------------------------------------------------------------
def _sample_text(value: object, limit: int = 50) -> str:
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return text[:limit]


def _stats(values: list[float]) -> dict[str, float] | None:
    if not values:
        return None
    avg = sum(values) / len(values)
    variance = sum((v - avg) ** 2 for v in values) / len(values)
    return {
        "min": round(min(values), 4),
        "max": round(max(values), 4),
        "avg": round(avg, 4),
        "stddev": round(math.sqrt(variance), 4),
    }


def _color_to_key(color) -> str | None:
    color_type = getattr(color, "type", None)
    if color_type == "rgb" and getattr(color, "rgb", None):
        return color.rgb
    if color_type == "indexed" and getattr(color, "indexed", None) is not None:
        return f"indexed:{color.indexed}"
    if color_type == "theme" and getattr(color, "theme", None) is not None:
        return f"theme:{color.theme}"
    if color_type == "auto":
        return "auto"
    return None


def _get_outline_level(dimension) -> int:
    return int(getattr(dimension, "outlineLevel", getattr(dimension, "outline_level", 0)) or 0)


def _count_cf_rules(ws) -> int:
    if hasattr(ws.conditional_formatting, "_cf_rules"):
        return sum(len(rules) for rules in ws.conditional_formatting._cf_rules.values())
    try:
        return len(ws.conditional_formatting)
    except Exception:
        return 0


def _get_print_area(ws) -> str | None:
    print_area = ws.print_area
    if not print_area:
        return None
    return str(print_area)


def _collect_dimension_stats(ws) -> tuple[dict[str, float] | None, dict[str, float] | None]:
    column_widths = [float(dim.width) for dim in ws.column_dimensions.values() if dim.width is not None]
    row_heights = [float(dim.height) for dim in ws.row_dimensions.values() if dim.height is not None]
    return _stats(column_widths), _stats(row_heights)


def _classify_number_format(pattern: str) -> str:
    lowered = pattern.lower()
    if "[$-411]" in lowered or "gg" in lowered:
        return "和暦/日付"
    if "%" in pattern:
        return "パーセント"
    if "¥" in pattern or "[$" in pattern:
        return "通貨"
    if "/" in pattern or "yy" in lowered or "dd" in lowered:
        return "日付"
    if "0" in pattern or "#" in pattern:
        return "数値"
    if pattern == "@":
        return "文字列"
    return "その他"


def _estimate_layout_type(sheets: list[SheetDetail]) -> str:
    candidates: set[str] = set()

    for sheet in sheets:
        col_stddev = None
        if sheet.column_width_stats:
            col_stddev = sheet.column_width_stats.get("stddev")

        if col_stddev is not None and col_stddev <= FORM_GRID_STDDEV_THRESHOLD and not sheet.grid_lines_visible:
            candidates.add("form_grid")
        if sheet.has_tables or sheet.max_row > LARGE_SHEET_ROW_THRESHOLD:
            candidates.add("data_table")
        if sheet.print_area:
            candidates.add("report_print")

    if len(candidates) > 1:
        return "mixed"
    if len(candidates) == 1:
        return next(iter(candidates))
    return "unknown"


def inspect_file(filepath: Path) -> XlsxInspection:
    inspection = XlsxInspection(path=str(filepath), has_vba=filepath.suffix.lower() == ".xlsm")

    try:
        wb = load_workbook(str(filepath), read_only=False, data_only=False)
    except Exception as e:
        inspection.error = f"xlsx 読み込み失敗: {e}"
        return inspection

    sheet_details: list[SheetDetail] = []
    named_ranges: list[NamedRangeInfo] = []

    try:
        for name, defined_name in wb.defined_names.items():
            scope = "workbook"
            local_sheet_id = getattr(defined_name, "localSheetId", None)
            if local_sheet_id is not None and 0 <= local_sheet_id < len(wb.worksheets):
                scope = wb.worksheets[local_sheet_id].title
            named_ranges.append(
                NamedRangeInfo(
                    name=name,
                    scope=scope,
                    refers_to=getattr(defined_name, "attr_text", "") or "",
                )
            )

        for index, ws in enumerate(wb.worksheets, 1):
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            total_cells = max_row * max_col
            concrete_cells = [cell for cell in getattr(ws, "_cells", {}).values() if not isinstance(cell, MergedCell)]

            merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            table_names = list(ws.tables.keys())
            image_count = len(getattr(ws, "_images", []))
            cf_rule_count = _count_cf_rules(ws)
            dv_count = len(getattr(ws.data_validations, "dataValidation", []))
            filter_ref = ws.auto_filter.ref if ws.auto_filter and ws.auto_filter.ref else None
            freeze_panes_ref = str(ws.freeze_panes) if ws.freeze_panes else None

            hidden_row_count = 0
            row_outline = False
            for dim in ws.row_dimensions.values():
                if dim.hidden:
                    hidden_row_count += 1
                if _get_outline_level(dim) > 0:
                    row_outline = True

            hidden_col_count = 0
            col_outline = False
            for dim in ws.column_dimensions.values():
                if dim.hidden:
                    hidden_col_count += 1
                if _get_outline_level(dim) > 0:
                    col_outline = True

            number_formats: Counter[str] = Counter()
            fill_colors: Counter[str] = Counter()
            non_empty_cells = sum(1 for cell in concrete_cells if cell.value is not None)
            comment_count = 0
            formula_count = 0
            comment_samples: list[str] = []

            format_row_limit = max_row
            if total_cells > 100000:
                format_row_limit = min(max_row, FORMAT_SAMPLE_ROW_LIMIT)

            for cell in concrete_cells:
                if cell.comment is not None:
                    comment_count += 1
                    if len(comment_samples) < COMMENT_SAMPLE_LIMIT:
                        comment_samples.append(f"{cell.coordinate}: {_sample_text(cell.comment.text)}")

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

                if cell.row <= format_row_limit:
                    number_format = cell.number_format or "General"
                    number_formats[number_format] += 1
                    fill_key = _color_to_key(cell.fill.fgColor)
                    if fill_key:
                        fill_colors[fill_key] += 1

            column_width_stats, row_height_stats = _collect_dimension_stats(ws)
            grid_lines_visible = ws.sheet_view.showGridLines
            if grid_lines_visible is None:
                grid_lines_visible = True

            detail = SheetDetail(
                name=ws.title,
                index=index,
                is_hidden=ws.sheet_state != "visible",
                is_protected=bool(ws.protection.sheet),
                max_row=max_row,
                max_column=max_col,
                non_empty_cells=non_empty_cells,
                total_cells=total_cells,
                empty_ratio=round(1 - (non_empty_cells / total_cells), 4) if total_cells > 0 else 0.0,
                has_merged_cells=bool(merged_ranges),
                merged_cell_count=len(merged_ranges),
                merged_cell_ranges=merged_ranges[:MERGED_SAMPLE_LIMIT],
                has_tables=bool(table_names),
                table_count=len(table_names),
                table_names=table_names,
                has_images=image_count > 0,
                image_count=image_count,
                has_comments=comment_count > 0,
                comment_count=comment_count,
                comment_samples=comment_samples,
                has_formulas=formula_count > 0,
                formula_count=formula_count,
                has_conditional_formatting=cf_rule_count > 0,
                cf_rule_count=cf_rule_count,
                has_data_validation=dv_count > 0,
                dv_count=dv_count,
                has_auto_filter=filter_ref is not None,
                filter_ref=filter_ref,
                has_row_outline=row_outline,
                has_col_outline=col_outline,
                has_hidden_rows=hidden_row_count > 0,
                hidden_row_count=hidden_row_count,
                has_hidden_cols=hidden_col_count > 0,
                hidden_col_count=hidden_col_count,
                has_freeze_panes=freeze_panes_ref is not None,
                freeze_panes_ref=freeze_panes_ref,
                print_area=_get_print_area(ws),
                print_title_rows=ws.print_title_rows or None,
                print_title_cols=ws.print_title_cols or None,
                number_format_patterns=dict(number_formats.most_common(NUMBER_FORMAT_TOP_N)),
                color_fill_patterns=dict(fill_colors.most_common(COLOR_PATTERN_TOP_N)),
                grid_lines_visible=bool(grid_lines_visible),
                column_width_stats=column_width_stats,
                row_height_stats=row_height_stats,
            )
            sheet_details.append(detail)

        inspection.sheet_count = len(sheet_details)
        inspection.hidden_sheet_count = sum(1 for sheet in sheet_details if sheet.is_hidden)
        inspection.sheets = [asdict(sheet) for sheet in sheet_details]
        inspection.named_ranges = [asdict(named_range) for named_range in named_ranges]
        inspection.total_merged_cells = sum(sheet.merged_cell_count for sheet in sheet_details)
        inspection.total_formulas = sum(sheet.formula_count for sheet in sheet_details)
        inspection.total_comments = sum(sheet.comment_count for sheet in sheet_details)
        inspection.total_images = sum(sheet.image_count for sheet in sheet_details)
        inspection.estimated_layout_type = _estimate_layout_type(sheet_details)
        return inspection
    finally:
        wb.close()


def build_text_report(inspections: list[XlsxInspection]) -> str:
    lines: list[str] = []
    ok_files = [insp for insp in inspections if insp.error is None]

    lines.append("=" * 60)
    lines.append("xlsx 詳細調査結果")
    lines.append("=" * 60)
    lines.append("")

    if not inspections:
        lines.append("対象ファイルなし")
        lines.append("=" * 60)
        lines.append("以上")
        return "\n".join(lines)

    total_sheets = sum(insp.sheet_count for insp in ok_files)
    total_hidden_sheets = sum(insp.hidden_sheet_count for insp in ok_files)
    total_protected_sheets = sum(
        1 for insp in ok_files for sheet in insp.sheets if sheet["is_protected"]
    )

    lines.append("■ 全体サマリー")
    lines.append("")
    lines.append(f"  ファイル数: {len(inspections)}件")
    lines.append(f"  シート数合計: {total_sheets}")
    lines.append(f"  非表示シート数: {total_hidden_sheets}")
    lines.append(f"  保護シート数: {total_protected_sheets}")
    lines.append("")

    lines.append("■ シート構造分析")
    lines.append("")
    large_sheets: list[tuple[str, dict]] = []
    sparse_sheets: list[tuple[str, dict]] = []
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["max_row"] > LARGE_SHEET_ROW_THRESHOLD or sheet["max_column"] > LARGE_SHEET_COL_THRESHOLD:
                large_sheets.append((insp.path, sheet))
            if sheet["empty_ratio"] >= HIGH_EMPTY_RATIO_THRESHOLD:
                sparse_sheets.append((insp.path, sheet))

    lines.append(f"  大きなシート（100行超 or 20列超）: {len(large_sheets)}件")
    for path, sheet in large_sheets[:10]:
        lines.append(
            f"    - {Path(path).name} / {sheet['name']}: "
            f"{sheet['max_row']}行×{sheet['max_column']}列 (空率{sheet['empty_ratio']:.0%})"
        )
    if len(large_sheets) > 10:
        lines.append(f"    - 他{len(large_sheets) - 10}件")

    lines.append(f"  空率が高いシート（80%以上）: {len(sparse_sheets)}件")
    for path, sheet in sparse_sheets[:10]:
        lines.append(
            f"    - {Path(path).name} / {sheet['name']}: "
            f"空率{sheet['empty_ratio']:.0%}、grid_lines_visible={sheet['grid_lines_visible']}"
        )
    if len(sparse_sheets) > 10:
        lines.append(f"    - 他{len(sparse_sheets) - 10}件")
    lines.append("")

    lines.append("■ 結合セル分析")
    lines.append("")
    merged_sheets = [sheet for insp in ok_files for sheet in insp.sheets if sheet["has_merged_cells"]]
    total_merged = sum(sheet["merged_cell_count"] for sheet in merged_sheets)
    lines.append(f"  結合セルを含むシート数: {len(merged_sheets)}")
    lines.append(f"  結合セル合計数: {total_merged}")
    sample_merged: list[str] = []
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["merged_cell_ranges"]:
                for merged_range in sheet["merged_cell_ranges"][:3]:
                    sample_merged.append(f"{Path(insp.path).name} / {sheet['name']} / {merged_range}")
    if sample_merged:
        lines.append("  結合パターンのサンプル:")
        for sample in sample_merged[:10]:
            lines.append(f"    - {sample}")
    lines.append("")

    lines.append("■ 数式分析")
    lines.append("")
    formula_sheets = [sheet for insp in ok_files for sheet in insp.sheets if sheet["has_formulas"]]
    lines.append(f"  数式を含むシート数: {len(formula_sheets)}")
    lines.append(f"  数式セル合計数: {sum(sheet['formula_count'] for sheet in formula_sheets)}")
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["has_formulas"]:
                lines.append(f"    - {Path(insp.path).name} / {sheet['name']}: {sheet['formula_count']}セル")
    lines.append("")

    lines.append("■ コメント・注釈")
    lines.append("")
    comment_sheets = [sheet for insp in ok_files for sheet in insp.sheets if sheet["has_comments"]]
    lines.append(f"  コメントを含むシート数: {len(comment_sheets)}")
    lines.append(f"  コメント合計数: {sum(sheet['comment_count'] for sheet in comment_sheets)}")
    comment_samples: list[str] = []
    for insp in ok_files:
        for sheet in insp.sheets:
            for sample in sheet["comment_samples"]:
                comment_samples.append(f"{Path(insp.path).name} / {sheet['name']} / {sample}")
    if comment_samples:
        lines.append("  サンプル:")
        for sample in comment_samples[:10]:
            lines.append(f"    - {sample}")
    lines.append("")

    lines.append("■ 書式パターン")
    lines.append("")
    number_formats: Counter[str] = Counter()
    color_patterns: Counter[str] = Counter()
    for insp in ok_files:
        for sheet in insp.sheets:
            number_formats.update(sheet["number_format_patterns"])
            color_patterns.update(sheet["color_fill_patterns"])

    if number_formats:
        lines.append("  number_format の出現パターン:")
        for pattern, count in number_formats.most_common(NUMBER_FORMAT_TOP_N):
            label = _classify_number_format(pattern)
            lines.append(f"    - {pattern} ({label}): {count}セル")
    else:
        lines.append("  number_format の出現パターン: なし")

    if color_patterns:
        lines.append("  色の使用パターン:")
        for color, count in color_patterns.most_common(COLOR_PATTERN_TOP_N):
            lines.append(f"    - {color}: {count}セル")
    else:
        lines.append("  色の使用パターン: なし")
    lines.append("")

    lines.append("■ データ保護")
    lines.append("")
    protected_sheets = []
    hidden_sheets = []
    dv_sheets = []
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["is_protected"]:
                protected_sheets.append(f"{Path(insp.path).name} / {sheet['name']}")
            if sheet["is_hidden"]:
                hidden_sheets.append(f"{Path(insp.path).name} / {sheet['name']}")
            if sheet["has_data_validation"]:
                dv_sheets.append(f"{Path(insp.path).name} / {sheet['name']}: {sheet['dv_count']}件")

    lines.append(f"  シート保護: {len(protected_sheets)}件")
    for item in protected_sheets[:10]:
        lines.append(f"    - {item}")
    lines.append(f"  非表示シート: {len(hidden_sheets)}件")
    for item in hidden_sheets[:10]:
        lines.append(f"    - {item}")

    if any(insp.named_ranges for insp in ok_files):
        lines.append("  名前定義:")
        named_range_samples = [
            f"{Path(insp.path).name} / {named_range['scope']} / {named_range['name']} -> {named_range['refers_to']}"
            for insp in ok_files
            for named_range in insp.named_ranges
        ]
        for item in named_range_samples[:10]:
            lines.append(f"    - {item}")
    else:
        lines.append("  名前定義: なし")

    if dv_sheets:
        lines.append("  入力規則を含むシート:")
        for item in dv_sheets[:10]:
            lines.append(f"    - {item}")
    else:
        lines.append("  入力規則を含むシート: なし")
    lines.append("")

    lines.append("■ 印刷設定")
    lines.append("")
    print_sheets = []
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["print_area"]:
                print_sheets.append(
                    f"{Path(insp.path).name} / {sheet['name']}: "
                    f"print_area={sheet['print_area']}, rows={sheet['print_title_rows']}, cols={sheet['print_title_cols']}"
                )
    if print_sheets:
        for item in print_sheets[:10]:
            lines.append(f"  - {item}")
    else:
        lines.append("  print_area が設定されたシート: なし")
    lines.append("")

    lines.append("■ レイアウト推定")
    lines.append("")
    layout_counter: Counter[str] = Counter(insp.estimated_layout_type for insp in ok_files)
    for layout, count in layout_counter.most_common():
        lines.append(f"  {layout}: {count}ファイル")
    lines.append("")

    lines.append("■ Excel Table オブジェクト")
    lines.append("")
    table_sheets = []
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["has_tables"]:
                table_names = ", ".join(sheet["table_names"])
                table_sheets.append(f"{Path(insp.path).name} / {sheet['name']}: {table_names}")
    lines.append(f"  Table を含むシート数: {len(table_sheets)}")
    for item in table_sheets[:10]:
        lines.append(f"    - {item}")
    lines.append("")

    lines.append("■ アウトライン・フィルタ")
    lines.append("")
    outline_sheets = []
    filter_sheets = []
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["has_row_outline"] or sheet["has_col_outline"]:
                outline_sheets.append(
                    f"{Path(insp.path).name} / {sheet['name']}: "
                    f"row_outline={sheet['has_row_outline']}, col_outline={sheet['has_col_outline']}"
                )
            if sheet["has_auto_filter"]:
                filter_sheets.append(f"{Path(insp.path).name} / {sheet['name']}: {sheet['filter_ref']}")
    lines.append(f"  行列グループ化を含むシート: {len(outline_sheets)}")
    for item in outline_sheets[:10]:
        lines.append(f"    - {item}")
    lines.append(f"  フィルタ設定を含むシート: {len(filter_sheets)}")
    for item in filter_sheets[:10]:
        lines.append(f"    - {item}")
    lines.append("")

    lines.append("■ 画像")
    lines.append("")
    image_sheets = [sheet for insp in ok_files for sheet in insp.sheets if sheet["has_images"]]
    lines.append(f"  画像を含むシート数: {len(image_sheets)}")
    lines.append(f"  画像合計数: {sum(sheet['image_count'] for sheet in image_sheets)}")
    for insp in ok_files:
        for sheet in insp.sheets:
            if sheet["has_images"]:
                lines.append(f"    - {Path(insp.path).name} / {sheet['name']}: {sheet['image_count']}件")
    lines.append("")

    if any(insp.error for insp in inspections):
        lines.append("■ 読み込みエラー")
        lines.append("")
        for insp in inspections:
            if insp.error:
                lines.append(f"  - {insp.path}: {insp.error}")
        lines.append("")

    lines.append("=" * 60)
    lines.append("以上")
    return "\n".join(lines)


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(errors="replace")
        except Exception:
            pass

    parser = argparse.ArgumentParser(
        description="xlsx 詳細調査: シート構造、結合セル、Table、画像、印刷設定などを調査する"
    )
    parser.add_argument("target", help=".xlsx / .xlsm ファイルまたはフォルダ")
    parser.add_argument(
        "--output",
        "-o",
        default="inspect_xlsx_report",
        help="出力ファイル名（拡張子なし）。.json と .txt を出力 (デフォルト: inspect_xlsx_report)",
    )
    args = parser.parse_args()

    target = Path(args.target)
    if target.is_file() and target.suffix.lower() in (".xlsx", ".xlsm"):
        xlsx_files = [target]
    elif target.is_dir():
        xlsx_files = sorted([*target.rglob("*.xlsx"), *target.rglob("*.xlsm")])
    else:
        print(f"エラー: {target} は .xlsx/.xlsm ファイルでもフォルダでもありません", file=sys.stderr)
        sys.exit(1)

    if not xlsx_files:
        print("対象の .xlsx / .xlsm ファイルが見つかりません", file=sys.stderr)
        sys.exit(1)

    print(f"調査開始: {len(xlsx_files)}ファイル")

    inspections: list[XlsxInspection] = []
    for i, fpath in enumerate(xlsx_files, 1):
        print(f"  [{i}/{len(xlsx_files)}] {fpath.name}")
        inspections.append(inspect_file(fpath))

    print("調査完了")

    json_path = Path(f"{args.output}.json")
    json_path.write_text(
        json.dumps([asdict(insp) for insp in inspections], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"JSON レポート出力: {json_path}")

    text_report = build_text_report(inspections)
    txt_path = Path(f"{args.output}.txt")
    txt_path.write_text(text_report, encoding="utf-8")
    print(f"テキストレポート出力: {txt_path}")

    print("")
    print(text_report)


if __name__ == "__main__":
    main()
