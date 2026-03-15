"""テストデータ .xlsx 生成スクリプト

openpyxl で作成可能なパターンのテストデータを生成する。

使い方:
  python tools/generate_excel_test_data.py [--output-dir input/]
  # デフォルト出力先: input/excel/

生成ファイル:
  1. many_tables.xlsx            - 複数シートに業務表を配置したブック
  2. multiple_tables_sheet.xlsx  - 1 シート内に複数表が混在するブック
  3. merged_cells.xlsx           - 縦結合・横結合・複合ヘッダを含むブック
  4. formulas_and_formats.xlsx   - 数式、表示形式、条件付き書式を含むブック
  5. comments_and_annotations.xlsx - コメント、色、リンク、入力規則を含むブック
  6. many_images.xlsx            - 画像を多数配置したブック
  7. mixed_complex.xlsx          - 複数パターンを混在させた総合ブック
  8. large_workbook.xlsx         - 大きめのシートを含むブック
  9. change_history.xlsx         - 変更履歴主体のブック
 10. excel_form_grid.xlsx        - Excel 方眼紙レイアウトを再現したブック
 11. approval_request.xlsx       - 稟議・申請書レイアウトのブック
 12. invoice_print_layout.xlsx   - 帳票印刷レイアウトのブック
 13. timesheet_calendar.xlsx     - 勤怠表・シフト表のブック
 14. ledger_with_sections.xlsx   - 台帳形式と小計行を含むブック
 15. wareki_and_normalization.xlsx - 和暦・全角半角混在を含むブック
 16. protected_master_validation.xlsx - 非表示マスタと保護を含むブック
 17. outline_and_filter.xlsx     - アウトラインとフィルタ設定を含むブック
"""

from __future__ import annotations

import argparse
import io
from datetime import date, timedelta
from functools import lru_cache
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:
    raise SystemExit("openpyxl が必要です。`pip install openpyxl` を実行してください。") from exc

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:
    raise SystemExit("Pillow が必要です。`pip install Pillow` を実行してください。") from exc


HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
SUBHEADER_FILL = PatternFill("solid", fgColor="E2F0D9")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
NG_FILL = PatternFill("solid", fgColor="FCE4D6")
THIN_SIDE = Side(style="thin", color="B7C9E2")
ALL_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)
JP_FONT_CANDIDATES = [
    Path("C:/Windows/Fonts/msgothic.ttc"),
    Path("C:/Windows/Fonts/YuGothM.ttc"),
    Path("C:/Windows/Fonts/meiryo.ttc"),
    Path("C:/Windows/Fonts/MSMINCHO.TTC"),
]
WAREKI_NUMBER_FORMAT = '[$-411]ggge"\u5e74"m"\u6708"d"\u65e5"'


def _style_cell(cell, *, bold: bool = False, fill=None, number_format: str | None = None) -> None:
    cell.font = Font(bold=bold)
    cell.fill = fill or PatternFill(fill_type=None)
    cell.border = ALL_BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if number_format:
        cell.number_format = number_format


def _autosize_columns(ws, *, max_width: int = 36, sample_rows: int = 200) -> None:
    row_limit = min(ws.max_row, sample_rows)
    for col_idx in range(1, ws.max_column + 1):
        width = 10
        for row_idx in range(1, row_limit + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _set_cell_value(
    cell,
    value: object,
    *,
    bold: bool = False,
    fill=None,
    alignment: Alignment | None = None,
    number_format: str | None = None,
) -> None:
    cell.value = value
    _style_cell(cell, bold=bold, fill=fill, number_format=number_format)
    if alignment is not None:
        cell.alignment = alignment


def _apply_grid(
    ws,
    *,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    fill=None,
    alignment: Alignment | None = None,
) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            _style_cell(cell, fill=fill)
            if alignment is not None:
                cell.alignment = alignment


def _merge_block(
    ws,
    cell_range: str,
    value: object,
    *,
    bold: bool = False,
    fill=None,
    alignment: Alignment | None = None,
    number_format: str | None = None,
) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    _apply_grid(
        ws,
        start_row=min_row,
        end_row=max_row,
        start_col=min_col,
        end_col=max_col,
        fill=fill,
        alignment=alignment,
    )
    _set_cell_value(
        ws.cell(row=min_row, column=min_col),
        value,
        bold=bold,
        fill=fill,
        alignment=alignment,
        number_format=number_format,
    )
    if min_row != max_row or min_col != max_col:
        ws.merge_cells(cell_range)


@lru_cache(maxsize=16)
def _load_japanese_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for font_path in JP_FONT_CANDIDATES:
        if not font_path.exists():
            continue
        try:
            return ImageFont.truetype(str(font_path), size)
        except OSError:
            continue
    return ImageFont.load_default()


def _draw_centered_label(
    draw: ImageDraw.ImageDraw,
    *,
    box: tuple[int, int, int, int],
    text: str,
    fill: str | tuple[int, int, int],
    font_size: int,
) -> None:
    font = _load_japanese_font(font_size)
    spacing = 4
    left, top, right, bottom = box
    bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=spacing, align="center")
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    x = left + max(0, (right - left - text_width) // 2)
    y = top + max(0, (bottom - top - text_height) // 2)
    draw.multiline_text((x, y), text, fill=fill, font=font, spacing=spacing, align="center")


def _format_japanese_era(value: date) -> str:
    eras = [
        ("令和", date(2019, 5, 1), 2018),
        ("平成", date(1989, 1, 8), 1988),
        ("昭和", date(1926, 12, 25), 1925),
        ("大正", date(1912, 7, 30), 1911),
    ]
    for era_name, start_date, base_year in eras:
        if value >= start_date:
            era_year = value.year - base_year
            year_label = "元" if era_year == 1 else str(era_year)
            return f"{era_name}{year_label}年{value.month}月{value.day}日"
    return f"{value.year}年{value.month}月{value.day}日"


def _write_table(
    ws,
    *,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list[object]],
    table_name: str,
    title: str | None = None,
) -> int:
    current_row = start_row
    if title:
        ws.cell(row=current_row, column=start_col, value=title)
        _style_cell(ws.cell(row=current_row, column=start_col), bold=True, fill=SUBHEADER_FILL)
        ws.merge_cells(
            start_row=current_row,
            start_column=start_col,
            end_row=current_row,
            end_column=start_col + len(headers) - 1,
        )
        current_row += 1

    for offset, header in enumerate(headers):
        cell = ws.cell(row=current_row, column=start_col + offset, value=header)
        _style_cell(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            cell = ws.cell(row=current_row + row_offset, column=start_col + col_offset, value=value)
            _style_cell(cell)

    end_row = current_row + len(rows)
    end_col = start_col + len(headers) - 1
    table = Table(
        displayName=table_name,
        ref=f"{get_column_letter(start_col)}{current_row}:{get_column_letter(end_col)}{end_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return end_row


def _make_labeled_image(label: str, size: tuple[int, int], color: tuple[int, int, int]) -> XLImage:
    img = Image.new("RGB", size, color)
    draw = ImageDraw.Draw(img)
    draw.rounded_rectangle((4, 4, size[0] - 4, size[1] - 4), radius=12, outline="white", width=3)
    _draw_centered_label(
        draw,
        box=(12, 12, size[0] - 12, size[1] - 12),
        text=label,
        fill="white",
        font_size=max(14, min(size) // 7),
    )
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    xl_image = XLImage(buffer)
    xl_image.width, xl_image.height = size
    return xl_image


def generate_many_tables(output_dir: Path) -> Path:
    wb = Workbook()
    sheet_specs = [
        (
            "ユーザー管理",
            ["ID", "氏名", "メール", "権限"],
            [
                ["001", "田中太郎", "tanaka@example.com", "管理者"],
                ["002", "鈴木花子", "suzuki@example.com", "一般"],
                ["003", "佐藤次郎", "sato@example.com", "閲覧者"],
                ["004", "高橋美咲", "takahashi@example.com", "一般"],
            ],
            "UserTable",
        ),
        (
            "エラーコード",
            ["コード", "種別", "メッセージ", "対処法"],
            [
                ["E001", "入力", "必須項目が未入力です", "項目を入力してください"],
                ["E002", "入力", "形式が不正です", "正しい形式で入力してください"],
                ["E003", "DB", "接続に失敗しました", "DB サーバーの状態を確認"],
                ["E004", "DB", "タイムアウト", "ネットワーク状態を確認"],
                ["E005", "認証", "トークンが無効です", "再ログインしてください"],
            ],
            "ErrorTable",
        ),
        (
            "画面一覧",
            ["画面ID", "画面名", "概要"],
            [
                ["SC001", "ログイン画面", "認証を行う"],
                ["SC002", "ダッシュボード", "集計情報を表示"],
                ["SC003", "ユーザー一覧", "ユーザーの検索・表示"],
                ["SC004", "ユーザー詳細", "ユーザー情報の編集"],
                ["SC005", "帳票出力", "PDF/Excel 出力"],
            ],
            "ScreenTable",
        ),
        (
            "API一覧",
            ["Method", "Path", "説明", "認証"],
            [
                ["GET", "/api/users", "ユーザー一覧取得", "要"],
                ["POST", "/api/users", "ユーザー作成", "要"],
                ["GET", "/api/users/{id}", "ユーザー詳細取得", "要"],
                ["PUT", "/api/users/{id}", "ユーザー更新", "要"],
                ["DELETE", "/api/users/{id}", "ユーザー削除", "要（管理者）"],
            ],
            "ApiTable",
        ),
        (
            "環境情報",
            ["環境", "サーバー", "DB", "URL"],
            [
                ["開発", "dev-app01", "dev-db01", "https://dev.example.com"],
                ["検証", "stg-app01", "stg-db01", "https://stg.example.com"],
                ["本番", "prd-app01/02", "prd-db01(M)/02(S)", "https://www.example.com"],
            ],
            "EnvironmentTable",
        ),
        (
            "設定値",
            ["パラメータ", "デフォルト値", "説明"],
            [
                ["MAX_RETRY", "3", "リトライ回数上限"],
                ["TIMEOUT_SEC", "30", "タイムアウト秒数"],
                ["BATCH_SIZE", "1000", "バッチ処理単位"],
                ["LOG_LEVEL", "INFO", "ログレベル"],
                ["CACHE_TTL", "3600", "キャッシュ有効秒数"],
            ],
            "ConfigTable",
        ),
    ]

    for idx, (title, headers, rows, table_name) in enumerate(sheet_specs):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = title
        ws.freeze_panes = "A2"
        _write_table(ws, start_row=1, start_col=1, headers=headers, rows=rows, table_name=table_name)
        note = ws.cell(row=len(rows) + 4, column=1, value=f"補足: {title} のテスト用データ。")
        _style_cell(note, bold=True, fill=NOTE_FILL)
        ws.merge_cells(start_row=len(rows) + 4, start_column=1, end_row=len(rows) + 4, end_column=len(headers))
        _autosize_columns(ws)

    path = output_dir / "many_tables.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_multiple_tables_sheet(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "複数表混在"
    ws["A1"] = "1 シート内に表が点在するパターン"
    _style_cell(ws["A1"], bold=True, fill=NOTE_FILL)
    ws.merge_cells("A1:H1")

    _write_table(
        ws,
        start_row=3,
        start_col=1,
        headers=["帳票ID", "帳票名", "出力形式"],
        rows=[
            ["RP001", "売上日報", "PDF"],
            ["RP002", "月次実績", "Excel"],
            ["RP003", "在庫一覧", "Excel"],
        ],
        table_name="ReportTable",
        title="帳票一覧",
    )

    _write_table(
        ws,
        start_row=3,
        start_col=6,
        headers=["バッチID", "処理名", "実行時刻", "担当"],
        rows=[
            ["B001", "日次集計", "02:00", "運用A"],
            ["B002", "月次締め", "01:00", "運用B"],
            ["B003", "データ連携", "06:00/18:00", "運用A"],
        ],
        table_name="BatchTable",
        title="バッチ一覧",
    )

    _write_table(
        ws,
        start_row=13,
        start_col=2,
        headers=["項目", "内容"],
        rows=[
            ["備考", "2 つの表の間に空白列がある。"],
            ["確認観点", "used_range が広がっても個別の表を認識できるか。"],
            ["注意点", "途中のメモ行や離れたセルも混ざる。"],
        ],
        table_name="MemoTable",
        title="補足メモ",
    )

    ws["J12"] = "孤立セル"
    _style_cell(ws["J12"], bold=True, fill=WARN_FILL)
    ws["J13"] = "右下に単独のメモを配置"
    _style_cell(ws["J13"])
    ws.freeze_panes = "A3"
    _autosize_columns(ws)

    ws2 = wb.create_sheet("右寄せ配置")
    ws2["C2"] = "右寄せ・下寄せの表配置"
    _style_cell(ws2["C2"], bold=True, fill=NOTE_FILL)
    ws2.merge_cells("C2:I2")
    _write_table(
        ws2,
        start_row=5,
        start_col=4,
        headers=["機能ID", "機能名", "状態"],
        rows=[
            ["F001", "受注登録", "稼働中"],
            ["F002", "出荷指示", "テスト中"],
            ["F003", "在庫照会", "企画中"],
        ],
        table_name="FunctionTable",
    )
    _write_table(
        ws2,
        start_row=18,
        start_col=10,
        headers=["キー", "値"],
        rows=[
            ["TOKEN_LIMIT", "8000"],
            ["MAX_RETRY", "3"],
            ["TIMEOUT", "30"],
        ],
        table_name="ParameterTable",
    )
    _autosize_columns(ws2)

    path = output_dir / "multiple_tables_sheet.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_merged_cells(output_dir: Path) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "階層構造"
    headers = ["大分類", "中分類", "項目", "値"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        _style_cell(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_rows = [
        ["入力系", "ファイル入力", "形式", "CSV"],
        ["", "", "文字コード", "UTF-8"],
        ["", "DB入力", "接続先", "PostgreSQL"],
        ["", "", "タイムアウト", "30秒"],
        ["出力系", "帳票", "形式", "PDF"],
        ["", "", "用紙サイズ", "A4"],
        ["", "ログ", "出力先", "/var/log/app"],
    ]
    for row_idx, row_values in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _style_cell(cell)

    ws.merge_cells("A2:A5")
    ws.merge_cells("B2:B3")
    ws.merge_cells("B4:B5")
    ws.merge_cells("A6:A8")
    ws.merge_cells("B6:B7")
    for ref in ("A2", "B2", "B4", "A6", "B6"):
        ws[ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws2 = wb.create_sheet("複合ヘッダ")
    for idx, header in enumerate(["コード", "名称", "入力", "", "出力", ""], start=1):
        cell = ws2.cell(row=1, column=idx, value=header)
        _style_cell(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.merge_cells("A1:A2")
    ws2.merge_cells("B1:B2")
    ws2.merge_cells("C1:D1")
    ws2.merge_cells("E1:F1")
    subheaders = ["ファイル", "DB", "画面", "帳票"]
    for idx, header in enumerate(subheaders, start=3):
        cell = ws2.cell(row=2, column=idx, value=header)
        _style_cell(cell, bold=True, fill=SUBHEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows2 = [
        ["F001", "受注処理", "CSV", "SELECT/INSERT", "一覧画面", "受注伝票"],
        ["F002", "出荷処理", "-", "SELECT/UPDATE", "詳細画面", "出荷指示書"],
        ["F003", "在庫照会", "-", "SELECT", "照会画面", "-"],
        ["F004", "月次集計", "CSV", "SELECT", "-", "月次レポート"],
    ]
    for row_idx, row_values in enumerate(rows2, start=3):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            _style_cell(cell)

    ws3 = wb.create_sheet("不規則結合")
    headers3 = ["品目", "数量", "単価", "金額"]
    for idx, header in enumerate(headers3, start=1):
        cell = ws3.cell(row=1, column=idx, value=header)
        _style_cell(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    rows3 = [
        ["サーバー", 2, 500000, 1000000],
        ["ストレージ", 4, 200000, 800000],
        ["ネットワーク機器", 1, 300000, 300000],
        ["小計", "", "", 2100000],
        ["消費税（10%）", "", "", 210000],
        ["合計", "", "", 2310000],
    ]
    for row_idx, row_values in enumerate(rows3, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            _style_cell(cell, number_format="#,##0")
    for ref in ("A5:C5", "A6:C6", "A7:C7"):
        ws3.merge_cells(ref)
    for ref in ("A5", "A6", "A7"):
        ws3[ref].alignment = Alignment(horizontal="right", vertical="center")

    for sheet in wb.worksheets:
        _autosize_columns(sheet)

    path = output_dir / "merged_cells.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_formulas_and_formats(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "売上集計"
    headers = ["日付", "部門", "担当", "数量", "単価", "売上", "消費税", "税込売上", "達成率", "判定"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        _style_cell(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    base_date = date(2026, 4, 1)
    departments = ["営業1課", "営業2課", "営業3課", "営業4課"]
    reps = ["田中", "鈴木", "佐藤", "高橋", "伊藤", "渡辺"]

    for row_idx in range(2, 14):
        ws.cell(row=row_idx, column=1, value=base_date + timedelta(days=row_idx - 2))
        ws.cell(row=row_idx, column=2, value=departments[(row_idx - 2) % len(departments)])
        ws.cell(row=row_idx, column=3, value=reps[(row_idx - 2) % len(reps)])
        ws.cell(row=row_idx, column=4, value=10 + (row_idx * 2))
        ws.cell(row=row_idx, column=5, value=12000 + (row_idx * 350))
        ws.cell(row=row_idx, column=6, value=f"=D{row_idx}*E{row_idx}")
        ws.cell(row=row_idx, column=7, value=f"=F{row_idx}*0.1")
        ws.cell(row=row_idx, column=8, value=f"=F{row_idx}+G{row_idx}")
        ws.cell(row=row_idx, column=9, value=f"=F{row_idx}/250000")
        ws.cell(row=row_idx, column=10, value=f'=IF(I{row_idx}>=1,"達成","未達")')
        for col_idx in range(1, 11):
            _style_cell(ws.cell(row=row_idx, column=col_idx))

    for row_idx in range(2, 14):
        ws.cell(row=row_idx, column=1).number_format = "yyyy/mm/dd"
        ws.cell(row=row_idx, column=4).number_format = "#,##0"
        for col_idx in (5, 6, 7, 8):
            ws.cell(row=row_idx, column=col_idx).number_format = '"¥"#,##0'
        ws.cell(row=row_idx, column=9).number_format = "0.0%"

    total_row = 15
    ws.cell(row=total_row, column=1, value="合計")
    _style_cell(ws.cell(row=total_row, column=1), bold=True, fill=SUBHEADER_FILL)
    for col_idx in range(4, 9):
        col_letter = get_column_letter(col_idx)
        ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}13)")
        _style_cell(ws.cell(row=total_row, column=col_idx), bold=True, fill=SUBHEADER_FILL)
    ws.cell(row=total_row, column=9, value="=AVERAGE(I2:I13)")
    _style_cell(ws.cell(row=total_row, column=9), bold=True, fill=SUBHEADER_FILL)
    ws.cell(row=total_row, column=9).number_format = "0.0%"

    ws.conditional_formatting.add(
        "I2:I13",
        CellIsRule(operator="lessThan", formula=["1"], stopIfTrue=True, fill=WARN_FILL),
    )
    ws.conditional_formatting.add(
        "I2:I13",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"], stopIfTrue=True, fill=OK_FILL),
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:J13"
    _autosize_columns(ws)

    ws2 = wb.create_sheet("KPI")
    headers2 = ["KPI", "目標", "実績", "差分", "達成率"]
    rows2 = [
        ["処理件数", 1500, 1620, "=C2-B2", "=C2/B2"],
        ["障害件数", 5, 3, "=C3-B3", "=C3/B3"],
        ["平均応答(ms)", 700, 640, "=C4-B4", "=C4/B4"],
        ["帳票出力件数", 420, 390, "=C5-B5", "=C5/B5"],
    ]
    _write_table(ws2, start_row=1, start_col=1, headers=headers2, rows=rows2, table_name="KpiTable")
    for row_idx in range(2, 6):
        for col_idx in (2, 3, 4):
            ws2.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        ws2.cell(row=row_idx, column=5).number_format = "0.0%"
    ws2["G2"] = "備考"
    _style_cell(ws2["G2"], bold=True, fill=NOTE_FILL)
    ws2["G3"] = "数式セルは Excel で開いたときに再計算される。"
    _style_cell(ws2["G3"])
    _autosize_columns(ws2)

    path = output_dir / "formulas_and_formats.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_comments_and_annotations(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "レビュー管理"
    headers = ["項目ID", "項目名", "状態", "ルール", "参照URL", "担当"]
    rows = [
        ["I001", "顧客ID", "要確認", "半角英数字10桁", "https://example.com/spec/customer", "田中"],
        ["I002", "契約日", "OK", "未来日不可", "https://example.com/spec/contract", "鈴木"],
        ["I003", "利用料金", "NG", "0以上、上限なし", "https://example.com/spec/fee", "佐藤"],
        ["I004", "ステータス", "要確認", "コード表と一致", "https://example.com/spec/status", "高橋"],
    ]
    _write_table(ws, start_row=1, start_col=1, headers=headers, rows=rows, table_name="ReviewTable")

    status_fills = {"OK": OK_FILL, "要確認": WARN_FILL, "NG": NG_FILL}
    for row_idx in range(2, 6):
        status = ws.cell(row=row_idx, column=3).value
        ws.cell(row=row_idx, column=3).fill = status_fills[status]
        ws.cell(row=row_idx, column=4).comment = Comment(
            f"{ws.cell(row=row_idx, column=2).value} の補足仕様。外部IFとの整合確認が必要。",
            "review-bot",
        )
        ws.cell(row=row_idx, column=5).hyperlink = ws.cell(row=row_idx, column=5).value
        ws.cell(row=row_idx, column=5).style = "Hyperlink"

    dv = DataValidation(type="list", formula1='"OK,要確認,NG"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("C2:C20")
    ws["H2"] = "内部メモ"
    _style_cell(ws["H2"], bold=True, fill=NOTE_FILL)
    ws["H3"] = "H 列は非表示。抽出時に hidden 列をどう扱うか確認用。"
    _style_cell(ws["H3"])
    ws.column_dimensions["H"].hidden = True
    ws.row_dimensions[9].hidden = True
    _autosize_columns(ws)

    ws2 = wb.create_sheet("色分け一覧")
    ws2["A1"] = "状態別の色分け"
    _style_cell(ws2["A1"], bold=True, fill=NOTE_FILL)
    ws2.merge_cells("A1:D1")
    palette_rows = [
        ["正常", "緑", "E2F0D9", "処理可能"],
        ["注意", "黄", "FFF2CC", "確認が必要"],
        ["異常", "赤", "FCE4D6", "修正が必要"],
    ]
    _write_table(
        ws2,
        start_row=3,
        start_col=1,
        headers=["状態", "表示名", "色コード", "説明"],
        rows=palette_rows,
        table_name="PaletteTable",
    )
    ws2["B6"].comment = Comment("色だけでなくラベルも保持したい。", "review-bot")
    _autosize_columns(ws2)

    path = output_dir / "comments_and_annotations.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_many_images(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "画面一覧"
    ws["A1"] = "画像が複数配置されたシート"
    _style_cell(ws["A1"], bold=True, fill=NOTE_FILL)
    ws.merge_cells("A1:H1")

    image_specs = [
        ("ログイン画面", "B3", (280, 150), (70, 130, 180)),
        ("ダッシュボード", "J3", (280, 150), (46, 139, 87)),
        ("ユーザー一覧", "B18", (280, 150), (220, 53, 69)),
        ("帳票出力", "J18", (280, 150), (147, 112, 219)),
    ]
    for title, anchor, size, color in image_specs:
        ws.add_image(_make_labeled_image(title, size, color), anchor)
        cell = ws[anchor]
        cell.value = title
        _style_cell(cell, bold=True, fill=SUBHEADER_FILL)

    for col in ("B", "J"):
        ws.column_dimensions[col].width = 22
    for row in (3, 18):
        ws.row_dimensions[row].height = 26
    ws["A32"] = "備考"
    _style_cell(ws["A32"], bold=True, fill=NOTE_FILL)
    ws["B32"] = "画像アンカーと近くのセルテキストを一緒に扱えるか確認用。"
    _style_cell(ws["B32"])
    _autosize_columns(ws)

    ws2 = wb.create_sheet("アイコン一覧")
    headers = ["種別", "説明", "アイコン"]
    rows = [
        ["ユーザー", "ユーザー関連操作", ""],
        ["設定", "システム設定", ""],
        ["警告", "エラー・警告表示", ""],
    ]
    _write_table(ws2, start_row=1, start_col=1, headers=headers, rows=rows, table_name="IconTable")
    icon_specs = [
        ("U", "C2", (48, 48), (70, 130, 180)),
        ("S", "C3", (48, 48), (46, 139, 87)),
        ("!", "C4", (48, 48), (220, 53, 69)),
    ]
    for label, anchor, size, color in icon_specs:
        ws2.add_image(_make_labeled_image(label, size, color), anchor)
    _autosize_columns(ws2)

    path = output_dir / "many_images.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_mixed_complex(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "概要"
    ws["A1"] = "機能仕様書（Excel 総合テスト）"
    _style_cell(ws["A1"], bold=True, fill=NOTE_FILL)
    ws.merge_cells("A1:F1")
    meta_rows = [
        ["システム名", "販売管理システム"],
        ["版数", "2.4"],
        ["作成日", date(2026, 3, 15)],
        ["担当", "業務設計チーム"],
    ]
    _write_table(
        ws,
        start_row=3,
        start_col=1,
        headers=["項目", "内容"],
        rows=meta_rows,
        table_name="MetaTable",
    )
    ws["B6"].number_format = "yyyy/mm/dd"
    ws["D3"] = "レビューコメント"
    _style_cell(ws["D3"], bold=True, fill=SUBHEADER_FILL)
    ws["D4"] = "シートごとに別の観点を持たせる。"
    _style_cell(ws["D4"])
    ws["D4"].comment = Comment("構造保持と Markdown 化の両方を見る。", "architect")
    _autosize_columns(ws)

    ws2 = wb.create_sheet("機能一覧")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["機能ID", "機能名", "種別", "状態", "備考"],
        rows=[
            ["F001", "ログイン", "画面", "稼働中", "認証基盤連携あり"],
            ["F002", "受注一覧", "画面", "稼働中", "検索条件が多い"],
            ["F003", "CSV 出力", "帳票", "テスト中", "列順固定"],
            ["F004", "日次集計", "バッチ", "開発中", "深夜実行"],
        ],
        table_name="FeatureTable",
    )
    dv = DataValidation(type="list", formula1='"画面,帳票,バッチ,API"', allow_blank=False)
    ws2.add_data_validation(dv)
    dv.add("C2:C20")
    ws2.freeze_panes = "A2"
    _autosize_columns(ws2)

    ws3 = wb.create_sheet("フロー")
    ws3["A1"] = "ワークフロー図イメージ"
    _style_cell(ws3["A1"], bold=True, fill=NOTE_FILL)
    ws3.merge_cells("A1:G1")
    ws3.add_image(_make_labeled_image("申請 -> 承認 -> 完了", (420, 120), (91, 155, 213)), "B3")
    ws3["B11"] = "図の横に補足メモ"
    _style_cell(ws3["B11"], bold=True, fill=SUBHEADER_FILL)
    ws3["C11"] = "画像だけでなく近接テキストも保持したい。"
    _style_cell(ws3["C11"])
    _autosize_columns(ws3)

    ws4 = wb.create_sheet("設定")
    headers4 = ["パラメータ", "現在値", "初期値", "判定"]
    rows4 = [
        ["TOKEN_LIMIT", 8000, 6000, '=IF(B2>=C2,"OK","要確認")'],
        ["MAX_RETRY", 3, 3, '=IF(B3=C3,"OK","要確認")'],
        ["TIMEOUT_SEC", 45, 30, '=IF(B4<=60,"OK","要確認")'],
        ["BATCH_SIZE", 1500, 1000, '=IF(B5>=C5,"OK","要確認")'],
    ]
    _write_table(ws4, start_row=1, start_col=1, headers=headers4, rows=rows4, table_name="SettingsTable")
    for row_idx in range(2, 6):
        status_cell = ws4.cell(row=row_idx, column=4)
        _style_cell(status_cell, fill=OK_FILL)
    ws4["A8"] = "複合ヘッダ"
    _style_cell(ws4["A8"], bold=True, fill=SUBHEADER_FILL)
    ws4.merge_cells("A8:D8")
    ws4["A9"] = "出力"
    ws4["C9"] = "入力"
    _style_cell(ws4["A9"], bold=True, fill=HEADER_FILL)
    _style_cell(ws4["C9"], bold=True, fill=HEADER_FILL)
    ws4.merge_cells("A9:B9")
    ws4.merge_cells("C9:D9")
    _autosize_columns(ws4)

    hidden = wb.create_sheet("マスタ")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "状態"
    hidden["A2"] = "稼働中"
    hidden["A3"] = "テスト中"
    hidden["A4"] = "開発中"

    path = output_dir / "mixed_complex.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_large_workbook(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "取引明細"
    headers = [
        "取引ID",
        "取引日",
        "顧客コード",
        "顧客名",
        "商品コード",
        "数量",
        "単価",
        "金額",
        "ステータス",
        "担当",
        "拠点",
        "備考",
    ]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        _style_cell(cell, bold=True, fill=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    statuses = ["受付", "処理中", "完了", "差戻し"]
    offices = ["東京", "大阪", "名古屋", "福岡"]
    reps = ["田中", "鈴木", "佐藤", "高橋", "伊藤", "渡辺"]

    for row_idx in range(2, 4002):
        qty = (row_idx % 9) + 1
        price = 1200 + ((row_idx * 17) % 850)
        ws.cell(row=row_idx, column=1, value=f"TX{row_idx:06d}")
        ws.cell(row=row_idx, column=2, value=date(2025, 1, 1) + timedelta(days=row_idx % 365))
        ws.cell(row=row_idx, column=3, value=f"C{(row_idx % 500):04d}")
        ws.cell(row=row_idx, column=4, value=f"顧客{row_idx % 500:04d}")
        ws.cell(row=row_idx, column=5, value=f"P{(row_idx % 900):05d}")
        ws.cell(row=row_idx, column=6, value=qty)
        ws.cell(row=row_idx, column=7, value=price)
        ws.cell(row=row_idx, column=8, value=qty * price)
        ws.cell(row=row_idx, column=9, value=statuses[row_idx % len(statuses)])
        ws.cell(row=row_idx, column=10, value=reps[row_idx % len(reps)])
        ws.cell(row=row_idx, column=11, value=offices[row_idx % len(offices)])
        ws.cell(
            row=row_idx,
            column=12,
            value=f"案件{row_idx:06d} の処理内容。関連資料ID=DOC-{row_idx % 120:03d}。",
        )

    for row_idx in range(2, 4002):
        ws.cell(row=row_idx, column=2).number_format = "yyyy/mm/dd"
        for col_idx in (6, 7, 8):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{ws.max_row}"
    width_overrides = {
        "A": 14,
        "B": 12,
        "C": 12,
        "D": 14,
        "E": 12,
        "F": 8,
        "G": 10,
        "H": 12,
        "I": 10,
        "J": 10,
        "K": 10,
        "L": 42,
    }
    for col, width in width_overrides.items():
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("月次集計")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["月", "取引件数", "売上金額", "備考"],
        rows=[
            ["2025-01", 330, 5120000, "四半期末調整あり"],
            ["2025-02", 305, 4890000, "通常運用"],
            ["2025-03", 352, 5450000, "キャンペーン実施"],
            ["2025-04", 311, 4980000, "通常運用"],
            ["2025-05", 327, 5210000, "大型案件計上"],
            ["2025-06", 339, 5330000, "通常運用"],
        ],
        table_name="MonthlySummaryTable",
    )
    for row_idx in range(2, 8):
        ws2.cell(row=row_idx, column=2).number_format = "#,##0"
        ws2.cell(row=row_idx, column=3).number_format = "#,##0"
    _autosize_columns(ws2)

    path = output_dir / "large_workbook.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_change_history(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "変更履歴"
    headers = ["ページ", "種別", "年　月", "記　　事"]
    rows = [
        ["全", "新規", "2024/10", "初版作成"],
        ["3", "追加", "2025/01", "コンポーネント設計追加"],
        ["5-8", "修正", "2025/03", "API 設計見直し"],
        ["10", "削除", "2025/06", "旧バッチ処理削除"],
        ["12-14", "修正", "2025/09", "Excel 出力仕様を更新"],
    ]
    _write_table(ws, start_row=1, start_col=1, headers=headers, rows=rows, table_name="ChangeHistoryTable")
    ws["A9"] = "変更履歴は全角スペースを含む見出しにしてある。"
    _style_cell(ws["A9"], bold=True, fill=NOTE_FILL)
    ws.merge_cells("A9:D9")
    _autosize_columns(ws)

    ws2 = wb.create_sheet("設計概要")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["項目", "内容"],
        rows=[
            ["文書名", "ドキュメント処理パイプライン詳細設計書"],
            ["アーキテクチャ", "マイクロサービス"],
            ["DB", "PostgreSQL"],
            ["帳票", "PDF / Excel"],
        ],
        table_name="SummaryTable",
    )
    ws2["A8"] = "本文と変更履歴が別シートに分かれるケース。"
    _style_cell(ws2["A8"], bold=True, fill=NOTE_FILL)
    ws2.merge_cells("A8:B8")
    _autosize_columns(ws2)

    path = output_dir / "change_history.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_excel_form_grid(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "業務変更申請"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85

    for col_idx in range(1, 17):
        ws.column_dimensions[get_column_letter(col_idx)].width = 4.2
    for row_idx in range(1, 41):
        ws.row_dimensions[row_idx].height = 18

    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left = Alignment(vertical="top", wrap_text=True)
    _apply_grid(ws, start_row=1, end_row=40, start_col=1, end_col=16)
    _merge_block(ws, "B2:O3", "業務変更申請書", bold=True, fill=HEADER_FILL, alignment=centered)

    label_ranges = [
        ("B5:D5", "申請番号"),
        ("H5:J5", "起票日"),
        ("B7:D7", "申請部署"),
        ("H7:J7", "申請者"),
        ("B9:D10", "件名"),
        ("B12:D17", "変更内容"),
        ("B19:D24", "影響範囲"),
        ("B26:D30", "対応方針"),
        ("B32:D36", "備考"),
    ]
    value_ranges = [
        ("E5:G5", "WF-2026-0412"),
        ("K5:O5", date(2026, 4, 12)),
        ("E7:G7", "営業企画部"),
        ("K7:O7", "田中 太郎"),
        ("E9:O10", "受注 CSV 取込レイアウト変更"),
        ("E12:O17", "取引先コードの桁数を 8 桁から 10 桁に拡張する。\n既存IFとの互換性維持のため旧列も当面残す。"),
        ("E19:O24", "受注登録、夜間バッチ、帳票出力、外部連携 API に影響。\nテスト観点は入力チェック、変換処理、既存データ互換。"),
        ("E26:O30", "2026/04 リリースで先行対応し、旧IFは 2026/09 に廃止する。"),
        ("E32:O36", "関連資料: 基本設計書 v2.4 / API 定義書 rev.7"),
    ]
    for cell_range, value in label_ranges:
        _merge_block(ws, cell_range, value, bold=True, fill=SUBHEADER_FILL, alignment=centered)
    for cell_range, value in value_ranges:
        _merge_block(
            ws,
            cell_range,
            value,
            fill=NOTE_FILL,
            alignment=top_left if "\n" in str(value) or len(str(value)) > 18 else centered,
            number_format="yyyy/mm/dd" if isinstance(value, date) else None,
        )

    _merge_block(ws, "B38:O39", "備考: 方眼紙レイアウトのため、空白セルにも罫線が存在する。", fill=WARN_FILL, alignment=centered)

    ws2 = wb.create_sheet("記入例")
    ws2.sheet_view.showGridLines = False
    for col_idx in range(1, 13):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 4.5
    for row_idx in range(1, 26):
        ws2.row_dimensions[row_idx].height = 18
    _apply_grid(ws2, start_row=1, end_row=25, start_col=1, end_col=12)
    _merge_block(ws2, "B2:K3", "Excel 方眼紙の記入例", bold=True, fill=HEADER_FILL, alignment=centered)
    _merge_block(ws2, "B5:C6", "項目", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws2, "D5:K6", "値", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    sample_rows = [
        ("画面名", "受注一括登録"),
        ("帳票名", "受注一覧表"),
        ("変更理由", "顧客コード体系変更への対応"),
        ("注意事項", "帳票レイアウトは列位置依存"),
    ]
    for idx, (label, value) in enumerate(sample_rows):
        start_row = 7 + (idx * 3)
        end_row = start_row + 1
        _merge_block(ws2, f"B{start_row}:C{end_row}", label, bold=True, fill=SUBHEADER_FILL, alignment=centered)
        _merge_block(ws2, f"D{start_row}:K{end_row}", value, fill=NOTE_FILL, alignment=top_left)

    path = output_dir / "excel_form_grid.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_approval_request(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "稟議申請"
    ws.sheet_view.showGridLines = False

    for col_idx in range(1, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5 if col_idx not in (1, 15) else 3
    for row_idx in range(1, 34):
        ws.row_dimensions[row_idx].height = 20

    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left = Alignment(vertical="top", wrap_text=True)
    _apply_grid(ws, start_row=1, end_row=33, start_col=1, end_col=15)
    _merge_block(ws, "B2:I3", "設備購入稟議書", bold=True, fill=HEADER_FILL, alignment=centered)
    _merge_block(ws, "B5:D5", "起案部署", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "E5:G5", "情報システム部", fill=NOTE_FILL, alignment=centered)
    _merge_block(ws, "H5:J5", "起案者", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "K5:N5", "鈴木 花子", fill=NOTE_FILL, alignment=centered)
    _merge_block(ws, "B7:D8", "件名", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "E7:N8", "帳票出力サーバー増設の件", fill=NOTE_FILL, alignment=centered)
    _merge_block(ws, "B10:D14", "起案理由", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "E10:N14", "月末処理時に Excel 帳票出力が集中し、現行構成では応答遅延が発生するため。", fill=NOTE_FILL, alignment=top_left)
    _merge_block(ws, "B16:D20", "投資効果", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "E16:N20", "出力待ち時間を平均 40% 短縮し、夜間バッチの完了時刻を 30 分前倒しする。", fill=NOTE_FILL, alignment=top_left)
    _merge_block(ws, "B22:D25", "チェック項目", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    checklist = [
        "□ 予算取得済み",
        "□ セキュリティ審査済み",
        "□ 運用引継ぎ計画あり",
        "□ ベンダ見積取得済み",
    ]
    for idx, text in enumerate(checklist, start=22):
        _merge_block(ws, f"E{idx}:N{idx}", text, fill=NOTE_FILL, alignment=Alignment(vertical="center"))
    _merge_block(ws, "B29:D31", "備考", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "E29:N31", "添付: 見積書、構成図、保守契約更新案", fill=NOTE_FILL, alignment=top_left)

    approval_boxes = [("J2:K3", "担当課長", "J3"), ("L2:M3", "部長", "L3"), ("N2:O3", "役員", "N3")]
    for cell_range, title, anchor in approval_boxes:
        _merge_block(ws, cell_range, title, bold=True, fill=SUBHEADER_FILL, alignment=centered)
        ws.add_image(_make_labeled_image("承認", (70, 70), (180, 52, 52)), anchor)

    ws2 = wb.create_sheet("添付一覧")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["添付ID", "資料名", "版数", "説明"],
        rows=[
            ["ATT-01", "ベンダ見積書", "v1.0", "ハードウェア増設費用"],
            ["ATT-02", "構成図", "v2.1", "現行 / 変更後の比較"],
            ["ATT-03", "運用計画", "v0.9", "保守体制・監視項目"],
        ],
        table_name="AttachmentTable",
    )
    _autosize_columns(ws2)

    path = output_dir / "approval_request.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_invoice_print_layout(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"
    ws.sheet_view.showGridLines = False

    widths = {"A": 6, "B": 16, "C": 9, "D": 9, "E": 12, "F": 12, "G": 12, "H": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_idx in range(1, 38):
        ws.row_dimensions[row_idx].height = 19

    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    top_left = Alignment(vertical="top", wrap_text=True)

    _merge_block(ws, "A1:H2", "請求書", bold=True, fill=HEADER_FILL, alignment=centered)
    ws.add_image(_make_labeled_image("Sample\nCorp", (120, 60), (55, 96, 146)), "A1")
    _merge_block(ws, "A4:C5", "請求先", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "D4:H5", "株式会社サンプル商事 御中", fill=NOTE_FILL, alignment=top_left)
    _merge_block(ws, "A7:B7", "請求番号", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "C7:D7", "INV-202604-015", fill=NOTE_FILL, alignment=centered)
    _merge_block(ws, "E7:F7", "発行日", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "G7:H7", date(2026, 4, 30), fill=NOTE_FILL, alignment=centered, number_format="yyyy/mm/dd")
    _merge_block(ws, "A9:H9", "下記のとおりご請求申し上げます。", fill=NOTE_FILL, alignment=centered)

    headers = ["No", "品名", "数量", "単位", "単価", "金額", "消費税", "備考"]
    for idx, header in enumerate(headers, start=1):
        _set_cell_value(ws.cell(row=11, column=idx), header, bold=True, fill=SUBHEADER_FILL, alignment=centered)
    items = [
        [1, "月次帳票出力ライセンス", 20, "本", 12000, "=C12*E12", "=F12*0.1", "4月分"],
        [2, "API 連携保守", 1, "式", 180000, "=C13*E13", "=F13*0.1", "定額"],
        [3, "追加サーバー利用料", 2, "台", 85000, "=C14*E14", "=F14*0.1", "従量課金"],
        [4, "帳票テンプレート改修", 1, "式", 240000, "=C15*E15", "=F15*0.1", "別紙参照"],
    ]
    for row_idx, row_values in enumerate(items, start=12):
        for col_idx, value in enumerate(row_values, start=1):
            _set_cell_value(ws.cell(row=row_idx, column=col_idx), value, alignment=centered if col_idx != 8 else top_left)
            if col_idx in (5, 6, 7):
                ws.cell(row=row_idx, column=col_idx).number_format = '"¥"#,##0'

    summary_rows = [("小計", "=SUM(F12:F15)"), ("消費税", "=SUM(G12:G15)"), ("合計", "=F18+F19")]
    for offset, (label, formula) in enumerate(summary_rows, start=18):
        _merge_block(ws, f"D{offset}:E{offset}", label, bold=True, fill=SUBHEADER_FILL, alignment=right)
        _merge_block(ws, f"F{offset}:G{offset}", formula, fill=NOTE_FILL, alignment=right)
        ws[f"F{offset}"].number_format = '"¥"#,##0'
    _merge_block(ws, "A23:H26", "支払条件: 月末締め翌月末支払い\n振込先: 三井住友銀行 東京中央支店 普通 1234567", fill=NOTE_FILL, alignment=top_left)
    _merge_block(ws, "A28:H33", "備考: 印刷時は A4 縦・1 ページに収まるよう縮小印刷を行う。明細が多い場合は別紙を使用する。", fill=WARN_FILL, alignment=top_left)

    ws.print_area = "A1:H33"
    ws.print_title_rows = "1:11"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.oddHeader.center.text = "請求書"
    ws.oddFooter.right.text = "Page &[Page] / &N"
    ws.row_breaks.append(Break(id=33))

    ws2 = wb.create_sheet("明細別紙")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["No", "品名", "数量", "単価", "金額", "備考"],
        rows=[
            [5, "運用問合せ対応", 12, 8000, 96000, "チケット対応"],
            [6, "帳票検証支援", 6, 15000, 90000, "テスト支援"],
            [7, "夜間リリース立会い", 2, 35000, 70000, "休日作業"],
        ],
        table_name="InvoiceAttachmentTable",
    )
    for row_idx in range(2, 5):
        for col_idx in (4, 5):
            ws2.cell(row=row_idx, column=col_idx).number_format = '"¥"#,##0'
    ws2.print_area = "A1:F10"
    _autosize_columns(ws2)

    path = output_dir / "invoice_print_layout.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_timesheet_calendar(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "2026年04月勤怠"
    ws.freeze_panes = "D5"

    centered = Alignment(horizontal="center", vertical="center")
    weekdays = ["月", "火", "水", "木", "金", "土", "日"]
    weekend_colors = {5: WARN_FILL, 6: NG_FILL}
    status_fill = {"出": OK_FILL, "在": OK_FILL, "有": NOTE_FILL, "半": WARN_FILL, "休": NG_FILL, "遅": WARN_FILL}

    _merge_block(ws, "A1:AK1", "勤務実績表 2026年4月", bold=True, fill=HEADER_FILL, alignment=centered)
    _merge_block(ws, "A3:A4", "社員番号", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "B3:B4", "氏名", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "C3:C4", "所属", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    for col_idx, header in enumerate(["出勤日数", "有休日数", "休日日数", "備考"], start=34):
        _merge_block(ws, f"{get_column_letter(col_idx)}3:{get_column_letter(col_idx)}4", header, bold=True, fill=SUBHEADER_FILL, alignment=centered)

    for day in range(1, 31):
        col_idx = 4 + day - 1
        current = date(2026, 4, day)
        fill = weekend_colors.get(current.weekday(), HEADER_FILL)
        _set_cell_value(ws.cell(row=3, column=col_idx), day, bold=True, fill=fill, alignment=centered)
        _set_cell_value(ws.cell(row=4, column=col_idx), weekdays[current.weekday()], bold=True, fill=fill, alignment=centered)
        ws.column_dimensions[get_column_letter(col_idx)].width = 4

    employees = [
        ("E001", "田中", "営業1課", ["出", "出", "出", "出", "出", "休", "休", "出", "出", "有", "有", "出", "出", "休", "休", "出", "出", "出", "遅", "出", "休", "休", "出", "出", "出", "出", "出", "休", "休", "出"]),
        ("E002", "鈴木", "営業2課", ["出", "在", "出", "出", "出", "休", "休", "出", "遅", "出", "半", "出", "出", "休", "休", "出", "在", "出", "出", "出", "休", "休", "有", "有", "出", "出", "出", "休", "休", "出"]),
        ("E003", "佐藤", "運用", ["出", "出", "出", "遅", "出", "休", "休", "在", "出", "出", "出", "出", "半", "休", "休", "出", "出", "出", "出", "出", "休", "休", "出", "在", "出", "出", "出", "休", "休", "出"]),
        ("E004", "高橋", "開発", ["在", "在", "出", "出", "出", "休", "休", "出", "出", "出", "出", "出", "出", "休", "休", "有", "有", "出", "出", "遅", "休", "休", "出", "出", "出", "半", "出", "休", "休", "出"]),
    ]
    for row_idx, (emp_id, name, dept, statuses) in enumerate(employees, start=5):
        _set_cell_value(ws.cell(row=row_idx, column=1), emp_id, fill=NOTE_FILL, alignment=centered)
        _set_cell_value(ws.cell(row=row_idx, column=2), name, fill=NOTE_FILL, alignment=centered)
        _set_cell_value(ws.cell(row=row_idx, column=3), dept, fill=NOTE_FILL, alignment=centered)
        for offset, status in enumerate(statuses, start=4):
            _set_cell_value(ws.cell(row=row_idx, column=offset), status, fill=status_fill.get(status, NOTE_FILL), alignment=centered)
        # 半休を有休日数へ含めることで、抽出側が数式の意味差分を読めるか確認しやすくする。
        ws.cell(row=row_idx, column=34, value=f'=COUNTIF(D{row_idx}:AG{row_idx},"出")+COUNTIF(D{row_idx}:AG{row_idx},"在")+COUNTIF(D{row_idx}:AG{row_idx},"遅")')
        ws.cell(row=row_idx, column=35, value=f'=COUNTIF(D{row_idx}:AG{row_idx},"有")+COUNTIF(D{row_idx}:AG{row_idx},"半")')
        ws.cell(row=row_idx, column=36, value=f'=COUNTIF(D{row_idx}:AG{row_idx},"休")')
        ws.cell(row=row_idx, column=37, value="勤怠締め済み")
        for col_idx in range(34, 38):
            _style_cell(ws.cell(row=row_idx, column=col_idx), fill=NOTE_FILL)
            ws.cell(row=row_idx, column=col_idx).alignment = centered if col_idx < 37 else Alignment(vertical="center")

    ws.row_dimensions[11].hidden = True
    _merge_block(ws, "A12:C12", "日別集計", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    for day in range(1, 31):
        col_idx = 4 + day - 1
        _set_cell_value(ws.cell(row=12, column=col_idx), f'=COUNTIF({get_column_letter(col_idx)}5:{get_column_letter(col_idx)}8,"出")+COUNTIF({get_column_letter(col_idx)}5:{get_column_letter(col_idx)}8,"在")', fill=NOTE_FILL, alignment=centered)

    ws2 = wb.create_sheet("シフト")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["日付", "早番", "遅番", "夜間待機", "備考"],
        rows=[
            [date(2026, 4, 1), "田中", "鈴木", "佐藤", "月初対応"],
            [date(2026, 4, 2), "高橋", "田中", "鈴木", "通常"],
            [date(2026, 4, 3), "佐藤", "高橋", "田中", "障害対応待機"],
            [date(2026, 4, 4), "鈴木", "佐藤", "高橋", "土曜体制"],
        ],
        table_name="ShiftTable",
    )
    for row_idx in range(2, 6):
        ws2.cell(row=row_idx, column=1).number_format = "yyyy/mm/dd"
    _autosize_columns(ws2)

    path = output_dir / "timesheet_calendar.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_ledger_with_sections(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "取引台帳"
    ws.freeze_panes = "A4"

    centered = Alignment(horizontal="center", vertical="center")
    top_left = Alignment(vertical="top", wrap_text=True)
    _merge_block(ws, "A1:H1", "業務取引台帳", bold=True, fill=HEADER_FILL, alignment=centered)

    _merge_block(ws, "A3:H3", "■ 売上データ", bold=True, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="left", vertical="center"))
    sales_headers = ["伝票No", "日付", "得意先", "案件名", "数量", "単価", "金額", "備考"]
    for idx, header in enumerate(sales_headers, start=1):
        _set_cell_value(ws.cell(row=4, column=idx), header, bold=True, fill=HEADER_FILL, alignment=centered)
    sales_rows = [
        ["S-001", date(2026, 4, 1), "東都商事", "受注一覧帳票", 120, 1500, 180000, "初回納品"],
        ["S-002", date(2026, 4, 2), "西日本物産", "API 連携", 1, 320000, 320000, ""],
        ["S-003", date(2026, 4, 5), "北海運輸", "CSV 取込改修", 1, 280000, 280000, "仕様変更あり"],
    ]
    for row_idx, row_values in enumerate(sales_rows, start=5):
        for col_idx, value in enumerate(row_values, start=1):
            _set_cell_value(ws.cell(row=row_idx, column=col_idx), value, alignment=centered if col_idx not in (4, 8) else top_left)
        ws.cell(row=row_idx, column=2).number_format = "yyyy/mm/dd"
        for col_idx in (6, 7):
            ws.cell(row=row_idx, column=col_idx).number_format = '"¥"#,##0'
    # 結合セルの左上だけに数式を置き、右側は空セルのまま残すパターンを再現する。
    _merge_block(ws, "A8:F8", "売上小計", bold=True, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="right", vertical="center"))
    _merge_block(ws, "G8:H8", "=SUM(G5:G7)", fill=NOTE_FILL, alignment=Alignment(horizontal="right", vertical="center"))
    ws["G8"].number_format = '"¥"#,##0'

    _merge_block(ws, "A10:H10", "■ 入金データ", bold=True, fill=SUBHEADER_FILL, alignment=Alignment(horizontal="left", vertical="center"))
    pay_headers = ["入金日", "得意先", "入金区分", "金額", "消込先", "状態", "担当", "メモ"]
    for idx, header in enumerate(pay_headers, start=1):
        _set_cell_value(ws.cell(row=11, column=idx), header, bold=True, fill=HEADER_FILL, alignment=centered)
    pay_rows = [
        [date(2026, 4, 10), "東都商事", "銀行振込", 180000, "S-001", "消込済", "田中", ""],
        [date(2026, 4, 12), "西日本物産", "銀行振込", 320000, "S-002", "消込済", "鈴木", ""],
        [date(2026, 4, 18), "北海運輸", "手形", 280000, "S-003", "確認中", "佐藤", "期日確認待ち"],
    ]
    for row_idx, row_values in enumerate(pay_rows, start=12):
        for col_idx, value in enumerate(row_values, start=1):
            _set_cell_value(ws.cell(row=row_idx, column=col_idx), value, alignment=centered if col_idx != 8 else top_left)
        ws.cell(row=row_idx, column=1).number_format = "yyyy/mm/dd"
        ws.cell(row=row_idx, column=4).number_format = '"¥"#,##0'
    _merge_block(ws, "A15:C15", "摘要", bold=True, fill=SUBHEADER_FILL, alignment=centered)
    _merge_block(ws, "D15:H15", "4月分は帳票改修と連携改修が混在。入金データは台帳形式で管理。", fill=NOTE_FILL, alignment=top_left)

    ws["J3"] = "問い合わせ履歴"
    _style_cell(ws["J3"], bold=True, fill=SUBHEADER_FILL)
    ws["J4"] = "2026/04/11"
    ws["K4"] = "手形サイト確認依頼"
    ws["J5"] = "2026/04/17"
    ws["K5"] = "請求書再送依頼"
    for ref in ("J4", "K4", "J5", "K5"):
        _style_cell(ws[ref])

    for col_idx in range(1, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14 if col_idx not in (4, 8, 11) else 24

    ws2 = wb.create_sheet("部門別メモ")
    _merge_block(ws2, "A1:F1", "台帳補足", bold=True, fill=HEADER_FILL, alignment=centered)
    notes = [
        ("営業", "売上データは案件単位で管理"),
        ("経理", "入金区分ごとに消込運用が異なる"),
        ("運用", "問い合わせ履歴は別システムとも二重管理"),
    ]
    for row_idx, (dept, note) in enumerate(notes, start=3):
        _set_cell_value(ws2.cell(row=row_idx, column=1), dept, bold=True, fill=SUBHEADER_FILL)
        _merge_block(ws2, f"B{row_idx}:F{row_idx}", note, fill=NOTE_FILL, alignment=top_left)
    _autosize_columns(ws2)

    path = output_dir / "ledger_with_sections.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_wareki_and_normalization(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "和暦"
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wareki_dates = [
        ("契約開始日", date(2024, 4, 1), "和暦表示"),
        ("契約終了日", date(2027, 3, 31), "元号の切替確認"),
        ("検収日", date(2019, 5, 1), "令和元年開始日"),
    ]

    _write_table(
        ws,
        start_row=1,
        start_col=1,
        headers=["項目", "和暦表示", "元日付", "表示形式", "備考"],
        rows=[
            [label, _format_japanese_era(raw_date), raw_date, WAREKI_NUMBER_FORMAT, note]
            for label, raw_date, note in wareki_dates
        ],
        table_name="WarekiTable",
    )
    for row_idx in range(2, 5):
        ws.cell(row=row_idx, column=2).alignment = centered
        ws.cell(row=row_idx, column=3).number_format = WAREKI_NUMBER_FORMAT
        ws.cell(row=row_idx, column=3).alignment = centered
    ws["G2"] = "注記"
    _style_cell(ws["G2"], bold=True, fill=SUBHEADER_FILL)
    ws["G3"] = "B列は期待表示を固定文字列で保持し、C列に元の日付と書式コードを残す。"
    _style_cell(ws["G3"], fill=NOTE_FILL)
    _autosize_columns(ws)

    ws2 = wb.create_sheet("正規化観点")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["種別", "値", "備考"],
        rows=[
            ["全角英数", "ＡＢＣ１２３", "半角正規化候補"],
            ["半角英数", "ABC123", "比較用"],
            ["全角スペース", "顧客　一覧", "見出し判定に影響"],
            ["機種依存文字", "㈱サンプル①№5", "Unicode 正規化観点"],
            ["かな混在", "ｶﾀｶﾅ / カタカナ / かな", "表記ゆれ"],
            ["郵便番号", "012-3456", "先頭ゼロ保持"],
            ["電話番号", "03-1234-5678", "書式保持"],
        ],
        table_name="NormalizationTable",
    )
    ws2["E2"] = "コード値"
    _style_cell(ws2["E2"], bold=True, fill=SUBHEADER_FILL)
    ws2["E3"] = 12
    _style_cell(ws2["E3"], fill=NOTE_FILL)
    ws2["E3"].number_format = "000000"
    ws2["F2"] = "和暦テキスト"
    _style_cell(ws2["F2"], bold=True, fill=SUBHEADER_FILL)
    ws2["F3"] = "令和6年4月1日"
    _style_cell(ws2["F3"], fill=NOTE_FILL)
    ws2["B5"].comment = Comment("㈱、①、№ の扱いを比較する。", "review-bot")
    _autosize_columns(ws2)

    path = output_dir / "wareki_and_normalization.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_protected_master_validation(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "入力シート"

    headers = ["申請ID", "件名", "状態", "部署", "区分", "担当", "期限", "進捗率", "備考"]
    _write_table(
        ws,
        start_row=1,
        start_col=1,
        headers=headers,
        rows=[
            ["RQ-001", "帳票出力改善", "起案", "情報システム部", "設備", "田中", date(2026, 4, 30), 0.4, "レビュー中"],
            ["RQ-002", "CSV 取込改修", "承認待ち", "営業企画部", "開発", "鈴木", date(2026, 5, 15), 0.2, "起案済み"],
            ["RQ-003", "夜間バッチ監視", "完了", "運用部", "運用", "佐藤", date(2026, 4, 20), 1.0, "反映済み"],
        ],
        table_name="ProtectedInputTable",
    )
    for row_idx in range(2, 5):
        ws.cell(row=row_idx, column=7).number_format = "yyyy/mm/dd"
        ws.cell(row=row_idx, column=8).number_format = "0%"
        for col_idx in range(2, 8):
            ws.cell(row=row_idx, column=col_idx).protection = Protection(locked=False)
            ws.cell(row=row_idx, column=col_idx).fill = NOTE_FILL

    master = wb.create_sheet("マスタ")
    master["A1"] = "状態"
    master["B1"] = "部署"
    master["C1"] = "区分"
    for idx, value in enumerate(["起案", "承認待ち", "差戻し", "完了"], start=2):
        master.cell(row=idx, column=1, value=value)
    for idx, value in enumerate(["情報システム部", "営業企画部", "運用部"], start=2):
        master.cell(row=idx, column=2, value=value)
    for idx, value in enumerate(["設備", "開発", "運用"], start=2):
        master.cell(row=idx, column=3, value=value)
    for col_idx in range(1, 4):
        for row_idx in range(1, 6):
            _style_cell(master.cell(row=row_idx, column=col_idx))
    master.sheet_state = "hidden"
    master.protection.enable()

    wb.defined_names.add(DefinedName("StatusList", attr_text="'マスタ'!$A$2:$A$5"))
    wb.defined_names.add(DefinedName("DivisionList", attr_text="'マスタ'!$B$2:$B$4"))
    wb.defined_names.add(DefinedName("CategoryList", attr_text="'マスタ'!$C$2:$C$4"))

    status_dv = DataValidation(type="list", formula1="=StatusList", allow_blank=False)
    division_dv = DataValidation(type="list", formula1="=DivisionList", allow_blank=False)
    category_dv = DataValidation(type="list", formula1="=CategoryList", allow_blank=False)
    for dv, cell_range in ((status_dv, "C2:C20"), (division_dv, "D2:D20"), (category_dv, "E2:E20")):
        ws.add_data_validation(dv)
        dv.add(cell_range)

    ws["K2"] = "入力セル"
    _style_cell(ws["K2"], bold=True, fill=SUBHEADER_FILL)
    ws["K3"] = "黄色セルのみ編集可"
    _style_cell(ws["K3"], fill=NOTE_FILL)
    ws["K5"] = "保護状態"
    _style_cell(ws["K5"], bold=True, fill=SUBHEADER_FILL)
    ws["L5"] = "シート保護有効"
    _style_cell(ws["L5"], fill=NOTE_FILL)
    ws.protection.enable()
    ws.protection.password = "test"
    ws.protection.autoFilter = False
    ws.protection.sort = False

    ws2 = wb.create_sheet("照会")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["申請ID", "状態", "担当", "備考"],
        rows=[
            ["=入力シート!A2", "=入力シート!C2", "=入力シート!F2", "=入力シート!I2"],
            ["=入力シート!A3", "=入力シート!C3", "=入力シート!F3", "=入力シート!I3"],
            ["=入力シート!A4", "=入力シート!C4", "=入力シート!F4", "=入力シート!I4"],
        ],
        table_name="ProtectedLookupTable",
    )
    _autosize_columns(ws)
    _autosize_columns(ws2)

    path = output_dir / "protected_master_validation.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def generate_outline_and_filter(output_dir: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "工程一覧"
    ws.freeze_panes = "A2"
    centered = Alignment(horizontal="center", vertical="center")

    headers = ["工程", "タスク", "状態", "担当", "開始", "終了", "見積h", "実績h", "備考"]
    for idx, header in enumerate(headers, start=1):
        _set_cell_value(ws.cell(row=1, column=idx), header, bold=True, fill=HEADER_FILL, alignment=centered)

    rows = [
        ["要件定義", "ヒアリング", "完了", "田中", date(2026, 4, 1), date(2026, 4, 2), 8, 7, ""],
        ["要件定義", "仕様整理", "完了", "鈴木", date(2026, 4, 3), date(2026, 4, 4), 10, 9, ""],
        ["基本設計", "画面設計", "進行中", "佐藤", date(2026, 4, 5), date(2026, 4, 8), 16, 12, "帳票方眼紙あり"],
        ["基本設計", "IF設計", "進行中", "高橋", date(2026, 4, 6), date(2026, 4, 9), 14, 8, ""],
        ["開発", "入力画面", "未着手", "伊藤", date(2026, 4, 10), date(2026, 4, 18), 24, 0, ""],
        ["開発", "帳票出力", "未着手", "渡辺", date(2026, 4, 10), date(2026, 4, 20), 32, 0, "Excel 帳票"],
        ["開発", "夜間バッチ", "未着手", "田中", date(2026, 4, 12), date(2026, 4, 22), 20, 0, ""],
        ["試験", "単体試験", "未着手", "鈴木", date(2026, 4, 23), date(2026, 4, 28), 18, 0, ""],
        ["試験", "結合試験", "未着手", "佐藤", date(2026, 4, 29), date(2026, 5, 10), 30, 0, ""],
    ]
    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            _set_cell_value(ws.cell(row=row_idx, column=col_idx), value, alignment=centered if col_idx != 9 else Alignment(vertical="center"))
        for col_idx in (5, 6):
            ws.cell(row=row_idx, column=col_idx).number_format = "yyyy/mm/dd"
        for col_idx in (7, 8):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"

    # アウトライン折りたたみ、フィルタ、手動非表示を意図的に混在させ、
    # 表示状態と used_range のズレを観測しやすくする。
    ws.row_dimensions.group(2, 3, outline_level=1, hidden=False)
    ws.row_dimensions.group(4, 5, outline_level=1, hidden=False)
    ws.row_dimensions.group(6, 8, outline_level=1, hidden=True)
    ws.column_dimensions.group("G", "H", hidden=True)
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.auto_filter.ref = "A1:I10"
    ws.auto_filter.add_filter_column(2, ["進行中", "未着手"])
    ws.auto_filter.add_sort_condition("A2:A10")
    ws.row_dimensions[2].hidden = True
    ws.row_dimensions[3].hidden = True

    ws2 = wb.create_sheet("月次集計")
    _write_table(
        ws2,
        start_row=1,
        start_col=1,
        headers=["月", "工程", "件数", "備考"],
        rows=[
            ["2026-04", "要件定義", 2, "完了"],
            ["2026-04", "基本設計", 2, "進行中"],
            ["2026-04", "開発", 3, "未着手"],
            ["2026-04", "試験", 2, "未着手"],
        ],
        table_name="OutlineSummaryTable",
    )
    ws2.row_dimensions.group(2, 5, outline_level=1, hidden=False)
    _autosize_columns(ws)
    _autosize_columns(ws2)

    path = output_dir / "outline_and_filter.xlsx"
    wb.save(path)
    print(f"  生成: {path}")
    return path


def _resolve_output_dir(base_dir: Path, leaf_name: str) -> Path:
    if base_dir.name.lower() == leaf_name.lower():
        return base_dir
    return base_dir / leaf_name


def main() -> None:
    parser = argparse.ArgumentParser(description="テストデータ .xlsx 生成")
    parser.add_argument(
        "--output-dir",
        "-o",
        type=Path,
        default=Path("input"),
        help="出力先ベースディレクトリ (default: input/ -> input/excel/)",
    )
    args = parser.parse_args()

    output_dir = _resolve_output_dir(args.output_dir, "excel")
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"テストデータ生成先: {output_dir}/")
    print()

    generators = [
        generate_many_tables,
        generate_multiple_tables_sheet,
        generate_merged_cells,
        generate_formulas_and_formats,
        generate_comments_and_annotations,
        generate_many_images,
        generate_mixed_complex,
        generate_large_workbook,
        generate_change_history,
        generate_excel_form_grid,
        generate_approval_request,
        generate_invoice_print_layout,
        generate_timesheet_calendar,
        generate_ledger_with_sections,
        generate_wareki_and_normalization,
        generate_protected_master_validation,
        generate_outline_and_filter,
    ]
    for generator in generators:
        generator(output_dir)

    print()
    print(f"完了: {len(generators)} ファイル生成しました。")


if __name__ == "__main__":
    main()
