"""Step2: Excel 構造抽出器

.xlsx ファイルからシート・表・コメントを抽出し、中間表現 (IntermediateDocument) を構築する。

設計方針:
  - openpyxl で確定的に抽出（LLM は使わない）
  - シート名は見出しとして保持する
  - シート全体を 1 表に潰さず、レイアウト上つながった領域ごとに抽出する
  - コメントはセルテキストに「※注: ...」として付記する
  - 数式はキャッシュ値優先、値が空なら数式文字列へフォールバックする
  - 非表示シートはスキップし、ログに記録する
  - 画像は存在のみ IMAGE 要素として記録する
"""

from __future__ import annotations

import time
from collections import deque
from logging import getLogger
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from src.config import PipelineConfig
from src.models.intermediate import (
    CellData,
    Confidence,
    DocumentElement,
    ElementType,
    ImageElement,
    IntermediateDocument,
)
from src.models.metadata import (
    ExtractedFileRecord,
    FileMetadata,
    ProcessStatus,
    StepResult,
)

logger = getLogger(__name__)


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def _build_merge_map(ws: Worksheet) -> dict[tuple[int, int], tuple[int, int, int, int]]:
    """結合セルのマップを構築する。"""
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    for merge_range in ws.merged_cells.ranges:
        bounds = (
            merge_range.min_row,
            merge_range.min_col,
            merge_range.max_row,
            merge_range.max_col,
        )
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = bounds
    return merge_map


def _formula_text(cell: Cell | MergedCell) -> str:
    if isinstance(cell, MergedCell):
        return ""
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        return value.strip()
    return ""


def _cell_has_meaningful_content(
    value_cell: Cell | MergedCell,
    formula_cell: Cell | MergedCell,
) -> bool:
    """セルが抽出対象となる実質データを持つか判定する。"""
    if isinstance(value_cell, MergedCell) and isinstance(formula_cell, MergedCell):
        return False

    if not isinstance(value_cell, MergedCell):
        value = value_cell.value
        if value is not None and str(value).strip():
            return True
        if value_cell.comment and value_cell.comment.text and value_cell.comment.text.strip():
            return True

    if not isinstance(formula_cell, MergedCell):
        value = formula_cell.value
        if value is not None:
            text = str(value).strip()
            if text:
                return True
        if formula_cell.comment and formula_cell.comment.text and formula_cell.comment.text.strip():
            return True

    return False


def _cell_text(
    value_cell: Cell | MergedCell,
    formula_cell: Cell | MergedCell,
) -> str:
    """セルのテキスト表現を取得する。

    - キャッシュ値があれば優先
    - キャッシュ値が空で数式なら数式文字列へフォールバック
    - コメントがあれば「※注: ...」を付記
    """
    if isinstance(value_cell, MergedCell):
        if isinstance(formula_cell, MergedCell):
            return ""

    text = ""
    if not isinstance(value_cell, MergedCell):
        value = value_cell.value
        if value is not None:
            text = str(value).strip()

    if not text and not isinstance(formula_cell, MergedCell):
        formula_text = _formula_text(formula_cell)
        if formula_text:
            text = formula_text
        elif formula_cell.value is not None:
            text = str(formula_cell.value).strip()

    comment_source = value_cell if not isinstance(value_cell, MergedCell) else formula_cell
    if not isinstance(comment_source, MergedCell):
        if comment_source.comment and comment_source.comment.text:
            comment_text = comment_source.comment.text.strip()
            if comment_text:
                text = f"{text} ※注: {comment_text}" if text else f"※注: {comment_text}"

    return text


def _build_occupied_positions(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]],
) -> set[tuple[int, int]]:
    """実データを持つ座標を 4 近傍連結判定用に展開する。"""
    occupied: set[tuple[int, int]] = set()
    seen_merges: set[tuple[int, int, int, int]] = set()

    max_row = max(ws_values.max_row or 0, ws_formulas.max_row or 0)
    max_col = max(ws_values.max_column or 0, ws_formulas.max_column or 0)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            value_cell = ws_values.cell(row=r, column=c)
            formula_cell = ws_formulas.cell(row=r, column=c)

            if (r, c) in merge_map:
                bounds = merge_map[(r, c)]
                min_r, min_c, max_r, max_c = bounds
                if (r, c) != (min_r, min_c) or bounds in seen_merges:
                    continue
                if not _cell_has_meaningful_content(
                    ws_values.cell(row=min_r, column=min_c),
                    ws_formulas.cell(row=min_r, column=min_c),
                ):
                    continue
                seen_merges.add(bounds)
                for rr in range(min_r, max_r + 1):
                    for cc in range(min_c, max_c + 1):
                        occupied.add((rr, cc))
                continue

            if _cell_has_meaningful_content(value_cell, formula_cell):
                occupied.add((r, c))

    return occupied


def _find_connected_bounds(
    occupied: set[tuple[int, int]],
) -> list[tuple[int, int, int, int]]:
    """4 近傍で連結した領域ごとの矩形 bounds を返す。"""
    remaining = set(occupied)
    components: list[tuple[int, int, int, int]] = []

    while remaining:
        start = remaining.pop()
        queue: deque[tuple[int, int]] = deque([start])
        min_r = max_r = start[0]
        min_c = max_c = start[1]

        while queue:
            r, c = queue.popleft()
            if r < min_r:
                min_r = r
            if r > max_r:
                max_r = r
            if c < min_c:
                min_c = c
            if c > max_c:
                max_c = c

            for nr, nc in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if (nr, nc) in remaining:
                    remaining.remove((nr, nc))
                    queue.append((nr, nc))

        components.append((min_r, min_c, max_r, max_c))

    components.sort(key=lambda b: (b[0], b[1], b[2], b[3]))
    return components


def _extract_region_table(
    ws_values: Worksheet,
    ws_formulas: Worksheet,
    bounds: tuple[int, int, int, int],
    merge_map: dict[tuple[int, int], tuple[int, int, int, int]],
) -> tuple[list[list[CellData]], bool]:
    """矩形領域を CellData の 2 次元リストに変換する。"""
    min_r, min_c, max_r, max_c = bounds
    rows: list[list[CellData]] = []
    seen_merges: set[tuple[int, int, int, int]] = set()
    has_merged = False

    for r in range(min_r, max_r + 1):
        row_data: list[CellData] = []
        for c in range(min_c, max_c + 1):
            value_cell = ws_values.cell(row=r, column=c)
            formula_cell = ws_formulas.cell(row=r, column=c)

            if (r, c) in merge_map:
                merge_bounds = merge_map[(r, c)]
                m_min_r, m_min_c, m_max_r, m_max_c = merge_bounds

                if not (min_r <= m_min_r <= max_r and min_c <= m_min_c <= max_c):
                    continue
                if (r, c) != (m_min_r, m_min_c) or merge_bounds in seen_merges:
                    continue

                seen_merges.add(merge_bounds)
                text = _cell_text(
                    ws_values.cell(row=m_min_r, column=m_min_c),
                    ws_formulas.cell(row=m_min_r, column=m_min_c),
                )
                if not text:
                    continue

                has_merged = True
                row_data.append(CellData(
                    text=text,
                    row=r - min_r,
                    col=c - min_c,
                    rowspan=m_max_r - m_min_r + 1,
                    colspan=m_max_c - m_min_c + 1,
                    is_header=(r == min_r),
                ))
                continue

            if not _cell_has_meaningful_content(value_cell, formula_cell):
                continue

            row_data.append(CellData(
                text=_cell_text(value_cell, formula_cell),
                row=r - min_r,
                col=c - min_c,
                rowspan=1,
                colspan=1,
                is_header=(r == min_r),
            ))

        if row_data:
            rows.append(row_data)

    return rows, has_merged


def _count_images(ws: Worksheet) -> int:
    """シート内の画像数を返す。"""
    try:
        return len(ws._images)
    except Exception:
        return 0


def _get_outline_level(dimension) -> int:
    return int(getattr(dimension, "outlineLevel", getattr(dimension, "outline_level", 0)) or 0)


def _sheet_notes(ws: Worksheet) -> list[str]:
    """フィルタ・保護など Excel 固有機能の要約を返す。"""
    notes: list[str] = []

    if ws.auto_filter and ws.auto_filter.ref:
        notes.append(f"[Excel機能] フィルタ範囲: {ws.auto_filter.ref}")

    if bool(ws.protection.sheet):
        notes.append("[Excel機能] シート保護: 有効")

    dv_count = len(getattr(ws.data_validations, "dataValidation", []))
    if dv_count > 0:
        notes.append(f"[Excel機能] 入力規則: {dv_count}件")

    has_row_outline = any(_get_outline_level(dim) > 0 for dim in ws.row_dimensions.values())
    has_col_outline = any(_get_outline_level(dim) > 0 for dim in ws.column_dimensions.values())
    if has_row_outline or has_col_outline:
        parts: list[str] = []
        if has_row_outline:
            parts.append("行")
        if has_col_outline:
            parts.append("列")
        notes.append(f"[Excel機能] アウトライン: {'/'.join(parts)}グループ化あり")

    hidden_rows = sum(1 for dim in ws.row_dimensions.values() if dim.hidden)
    hidden_cols = sum(1 for dim in ws.column_dimensions.values() if dim.hidden)
    if hidden_rows > 0 or hidden_cols > 0:
        parts: list[str] = []
        if hidden_rows > 0:
            parts.append(f"非表示行={hidden_rows}")
        if hidden_cols > 0:
            parts.append(f"非表示列={hidden_cols}")
        notes.append(f"[Excel機能] {' / '.join(parts)}")

    return notes


# ---------------------------------------------------------------------------
# メイン抽出関数
# ---------------------------------------------------------------------------

def extract_xlsx(
    xlsx_path: Path,
    source_path: str,
    source_ext: str,
    config: PipelineConfig,
) -> tuple[ExtractedFileRecord, StepResult]:
    """1つの .xlsx ファイルから中間表現を抽出する。"""
    t0 = time.perf_counter()

    try:
        wb_values = load_workbook(str(xlsx_path), data_only=True, read_only=False)
        wb_formulas = load_workbook(str(xlsx_path), data_only=False, read_only=False)
    except Exception as e:
        elapsed = time.perf_counter() - t0
        result = StepResult(
            file_path=source_path,
            step="extract",
            status=ProcessStatus.ERROR,
            message=str(e),
            duration_sec=round(elapsed, 2),
        )
        meta = FileMetadata(source_path=source_path, source_ext=source_ext)
        record = ExtractedFileRecord(metadata=meta, document={})
        return record, result

    intermediate = IntermediateDocument()
    source_idx = 0
    total_sheets = 0
    hidden_sheets = 0
    total_tables = 0
    total_images = 0
    total_merged = 0
    warnings: list[str] = []

    try:
        for ws_values in wb_values.worksheets:
            if ws_values.sheet_state != "visible":
                hidden_sheets += 1
                logger.debug("非表示シートをスキップ: %s / %s", source_path, ws_values.title)
                continue

            ws_formulas = wb_formulas[ws_values.title]
            total_sheets += 1

            intermediate.add_heading(
                level=2,
                text=ws_values.title,
                detection_method="sheet_name",
                source_index=source_idx,
            )
            source_idx += 1

            img_count = _count_images(ws_values)
            if img_count > 0:
                total_images += img_count
                for _ in range(img_count):
                    intermediate.elements.append(DocumentElement(
                        type=ElementType.IMAGE,
                        content=ImageElement(
                            alt_text="",
                            description=f"画像 ({ws_values.title} 内)",
                        ),
                        source_index=source_idx,
                    ))
                    source_idx += 1

            merge_map = _build_merge_map(ws_formulas)
            occupied = _build_occupied_positions(ws_values, ws_formulas, merge_map)

            if not occupied:
                intermediate.add_paragraph("(空のシート)", source_index=source_idx)
                source_idx += 1
            else:
                component_bounds = _find_connected_bounds(occupied)
                total_tables += len(component_bounds)

                if ws_values.merged_cells.ranges:
                    total_merged += 1

                for bounds in component_bounds:
                    rows, has_merged = _extract_region_table(
                        ws_values, ws_formulas, bounds, merge_map,
                    )
                    if not rows:
                        continue

                    num_rows = len(rows)
                    num_cols = max(
                        (max(cell.col + cell.colspan for cell in row) for row in rows),
                        default=0,
                    )
                    if (
                        num_rows > config.excel_large_sheet_rows
                        or num_cols > config.excel_large_sheet_cols
                    ):
                        warnings.append(
                            f"large_sheet:{ws_values.title}({num_rows}r×{num_cols}c)"
                        )

                    confidence = Confidence.MEDIUM if has_merged else Confidence.HIGH
                    min_r, min_c, max_r, max_c = bounds
                    intermediate.add_table(
                        rows=rows,
                        caption="",
                        has_merged_cells=has_merged,
                        confidence=confidence,
                        fallback_reason="",
                        source_row_start=min_r,
                        source_col_start=min_c,
                        source_row_end=max_r,
                        source_col_end=max_c,
                        source_index=source_idx,
                    )
                    source_idx += 1

            for note in _sheet_notes(ws_values):
                intermediate.add_paragraph(note, source_index=source_idx)
                source_idx += 1
    finally:
        wb_values.close()
        wb_formulas.close()

    meta = FileMetadata(
        source_path=source_path,
        source_ext=source_ext,
        source_size_bytes=xlsx_path.stat().st_size,
        normalized_from=source_ext if source_ext != ".xlsx" else "",
        doc_role_guess="data_sheet",
    )

    record = ExtractedFileRecord(
        metadata=meta,
        document=intermediate.to_dict(),
    )

    elapsed = time.perf_counter() - t0

    if hidden_sheets > 0:
        warnings.append(f"hidden_sheets={hidden_sheets}")
    if total_images > 0:
        warnings.append(f"images={total_images}")
    if total_merged > 0:
        warnings.append(f"merged_cell_sheets={total_merged}")

    status = ProcessStatus.SUCCESS
    msg = (
        f"sheets={total_sheets}, tables={total_tables}, "
        f"elements={len(intermediate.elements)}"
    )
    if warnings:
        status = ProcessStatus.WARNING
        msg += ", " + ", ".join(warnings)

    result = StepResult(
        file_path=source_path,
        step="extract",
        status=status,
        message=msg,
        duration_sec=round(elapsed, 2),
    )
    logger.info("抽出完了: %s (%s)", source_path, msg)

    return record, result
